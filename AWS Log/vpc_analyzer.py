# -*- coding: utf-8 -*-
"""
VPC Flow Log Analyzer
Analyzes VPC Flow Log data and generates Excel and HTML reports
"""

import pandas as pd
import numpy as np
import os
import glob
import json
from datetime import datetime, timezone, timedelta
import sys
import codecs
import locale

# Windows Korean encoding issue resolution
if sys.platform.startswith('win'):
    # Environment variable setting
    os.environ['PYTHONIOENCODING'] = 'utf-8'
    
    # stdout/stderr encoding setting
    try:
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8')
            sys.stderr.reconfigure(encoding='utf-8')
        else:
            sys.stdout = codecs.getwriter('utf-8')(sys.stdout.detach())
            sys.stderr = codecs.getwriter('utf-8')(sys.stderr.detach())
    except:
        # Fallback for environments where detach is not available
        pass
    
    # Locale setting
    try:
        locale.setlocale(locale.LC_ALL, 'ko_KR.UTF-8')
    except:
        try:
            locale.setlocale(locale.LC_ALL, 'Korean_Korea.949')
        except:
            pass


class VPCFlowLogAnalyzer:
    def __init__(self, csv_file):
        """Initialize VPC Flow Log Analyzer"""
        self.csv_file = csv_file
        self.df = None
        self.load_data()
    
    def load_data(self):
        """Load VPC Flow Log data from CSV file"""
        try:
            print(f"Loading VPC Flow Log data from: {self.csv_file}")
            self.df = pd.read_csv(self.csv_file)
            print(f"Loaded data: {len(self.df)} flow records")
            
            # Convert timestamp columns to datetime (handle new column names)
            if 'window_start_(UTC+0)' in self.df.columns:
                # Handle new column format: YYYY-MM-DD HH:MM:SS
                self.df['start'] = pd.to_datetime(self.df['window_start_(UTC+0)'], errors='coerce')
                # Remove timezone info for Excel compatibility
                self.df['start'] = self.df['start'].dt.tz_localize(None)
            elif 'window_start' in self.df.columns:
                # Fallback for old format
                self.df['start'] = pd.to_datetime(
                    self.df['window_start'].str.replace(' (UTC+0)', '', regex=False), 
                    errors='coerce'
                )
                # Remove timezone info for Excel compatibility
                self.df['start'] = self.df['start'].dt.tz_localize(None)
            
            if 'window_end_(UTC+0)' in self.df.columns:
                # Handle new column format: YYYY-MM-DD HH:MM:SS
                self.df['end'] = pd.to_datetime(self.df['window_end_(UTC+0)'], errors='coerce')
                # Remove timezone info for Excel compatibility
                self.df['end'] = self.df['end'].dt.tz_localize(None)
            elif 'window_end' in self.df.columns:
                # Fallback for old format
                self.df['end'] = pd.to_datetime(
                    self.df['window_end'].str.replace(' (UTC+0)', '', regex=False), 
                    errors='coerce'
                )
                # Remove timezone info for Excel compatibility
                self.df['end'] = self.df['end'].dt.tz_localize(None)
            
            # Convert numeric columns (exclude timestamp columns)
            numeric_columns = ['srcport', 'dstport', 'protocol', 'packets', 'bytes']
            for col in numeric_columns:
                if col in self.df.columns:
                    self.df[col] = pd.to_numeric(self.df[col], errors='coerce')
            
        except Exception as e:
            print(f"Error loading data: {e}")
            self.df = pd.DataFrame()
    
    def analyze_top_ips_by_bytes(self, top_n=20):
        """1. Top 20 IPs by total bytes (srcIP, dstIP)"""
        print(f"\n1. Analyzing top {top_n} IPs by bytes...")
        
        if self.df.empty or 'bytes' not in self.df.columns:
            return pd.DataFrame({'IP': ['No data'], 'Total Bytes': [0], 'Type': ['N/A']})
        
        # Source IP statistics
        src_stats = self.df.groupby('srcaddr').agg({
            'bytes': 'sum',
            'packets': 'sum'
        }).reset_index()
        src_stats['Type'] = 'Source IP'
        src_stats = src_stats.rename(columns={'srcaddr': 'IP', 'bytes': 'Total Bytes'})
        
        # Destination IP statistics
        dst_stats = self.df.groupby('dstaddr').agg({
            'bytes': 'sum',
            'packets': 'sum'
        }).reset_index()
        dst_stats['Type'] = 'Destination IP'
        dst_stats = dst_stats.rename(columns={'dstaddr': 'IP', 'bytes': 'Total Bytes'})
        
        # Combine and sort
        combined_stats = pd.concat([src_stats, dst_stats], ignore_index=True)
        combined_stats = combined_stats.sort_values('Total Bytes', ascending=False).head(top_n)
        
        return combined_stats[['IP', 'Total Bytes', 'Type', 'packets']].rename(columns={'packets': 'Total Packets'})
    
    def analyze_top_ports_protocols(self, top_n=20):
        """2. Top ports and protocols"""
        print(f"\n2. Analyzing top {top_n} ports and protocols...")
        
        if self.df.empty:
            return pd.DataFrame({'Port': [0], 'Protocol': [0], 'Flow Count': [0], 'Total Bytes': [0]})
        
        # Port statistics
        port_stats = self.df.groupby(['dstport', 'protocol']).agg({
            'bytes': 'sum',
            'packets': 'sum'
        }).reset_index()
        port_stats['Flow Count'] = self.df.groupby(['dstport', 'protocol']).size().values
        port_stats = port_stats.sort_values('bytes', ascending=False).head(top_n)
        
        # Map protocol numbers to names
        protocol_map = {1: 'ICMP', 6: 'TCP', 17: 'UDP', 47: 'GRE', 50: 'ESP', 51: 'AH'}
        port_stats['Protocol'] = port_stats['protocol'].map(protocol_map).fillna(port_stats['protocol'])
        
        return port_stats[['dstport', 'Protocol', 'Flow Count', 'bytes']].rename(columns={
            'dstport': 'Port',
            'bytes': 'Total Bytes'
        })
    
    def analyze_nightshift_remote_access(self):
        """3. Nightshift remote access events (RDP, SSH, SCP, FTP/SFTP)"""
        print(f"\n3. Analyzing nightshift remote access events...")
        
        if self.df.empty or 'start' not in self.df.columns:
            return pd.DataFrame({'Source IP': ['No data'], 'Destination IP': ['No data'], 'Port': [0], 'Protocol': ['N/A'], 'Service': ['N/A'], 'Bytes': [0]})
        
        # Convert start time to datetime and extract hour
        self.df['start'] = pd.to_datetime(self.df['start'], errors='coerce')
        self.df['hour'] = self.df['start'].dt.hour
        
        # Filter nightshift hours (22:00-06:00)
        nightshift_mask = (self.df['hour'] >= 22) | (self.df['hour'] < 6)
        nightshift_flows = self.df[nightshift_mask]
        
        if len(nightshift_flows) == 0:
            return pd.DataFrame({'Source IP': ['No nightshift events'], 'Destination IP': ['N/A'], 'Port': [0], 'Protocol': ['N/A'], 'Service': ['N/A'], 'Bytes': [0]})
        
        # Define remote access ports (DFIR focused - high-risk protocols only)
        remote_ports = {
            # SSH, SCP, SFTP (all use port 22)
            22: 'SSH/SCP/SFTP',
            2222: 'SSH Alt',
            22222: 'SSH Alt',
            
            # Telnet (unencrypted - high risk)
            23: 'Telnet',
            9923: 'Telnet over TLS',
            
            # FTP (file transfer protocols)
            21: 'FTP',
            20: 'FTP Data',
            990: 'FTP over TLS',
            989: 'FTP over TLS Data',
            2121: 'FTP Alt',
            
            # RDP (Remote Desktop) - Windows remote access
            3389: 'RDP',
            3388: 'RDP Alt',
            3387: 'RDP Alt',
            3386: 'RDP Alt',
            3385: 'RDP Alt',
            3384: 'RDP Alt',
            3383: 'RDP Alt',
            3382: 'RDP Alt',
            3381: 'RDP Alt',
            3380: 'RDP Alt',
            
            # VNC (Virtual Network Computing) - cross-platform remote access
            5900: 'VNC',
            5901: 'VNC Display 1',
            5902: 'VNC Display 2',
            5903: 'VNC Display 3',
            5904: 'VNC Display 4',
            5905: 'VNC Display 5',
            5906: 'VNC Display 6',
            
            # TeamViewer - commercial remote access
            5938: 'TeamViewer',
            5939: 'TeamViewer',
            5940: 'TeamViewer',
            
            # Remote Management (Windows)
            5985: 'WinRM HTTP',
            5986: 'WinRM HTTPS',
            
            # Remote Administration (Windows)
            135: 'RPC Endpoint Mapper',
            139: 'NetBIOS Session',
            445: 'SMB/CIFS',
            
            # Remote Shell (Unix/Linux)
            512: 'Rexec',
            513: 'Rlogin',
            514: 'RSH',
            515: 'LPD',
            
            # High-risk remote access ports
            5631: 'PC Anywhere',
            5632: 'PC Anywhere',
            
            # Remote desktop alternatives
            4000: 'VNC Alt',
            4001: 'VNC Alt',
            4002: 'VNC Alt',
            4003: 'VNC Alt',
            4004: 'VNC Alt',
            4005: 'VNC Alt',
            4006: 'VNC Alt'
        }
        
        # Filter for remote access ports
        remote_access_mask = nightshift_flows['dstport'].isin(remote_ports.keys())
        remote_flows = nightshift_flows[remote_access_mask].copy()
        
        if len(remote_flows) == 0:
            return pd.DataFrame({'Source IP': ['No remote access events'], 'Destination IP': ['N/A'], 'Port': [0], 'Protocol': ['N/A'], 'Service': ['N/A'], 'Total Bytes': [0]})
        
        # Add service name
        remote_flows['Service'] = remote_flows['dstport'].map(remote_ports)
        
        # Group by connection and sum bytes
        connection_stats = remote_flows.groupby(['srcaddr', 'dstaddr', 'dstport', 'protocol', 'Service']).agg({
            'bytes': 'sum',
            'packets': 'sum'
        }).reset_index()
        
        # Map protocol numbers to names
        protocol_map = {1: 'ICMP', 6: 'TCP', 17: 'UDP', 47: 'GRE', 50: 'ESP', 51: 'AH'}
        connection_stats['Protocol'] = connection_stats['protocol'].map(protocol_map).fillna(connection_stats['protocol'])
        
        result_df = connection_stats[['srcaddr', 'dstaddr', 'dstport', 'Protocol', 'Service', 'bytes', 'packets']].copy()
        result_df = result_df.rename(columns={
            'srcaddr': 'Source IP',
            'dstaddr': 'Destination IP',
            'dstport': 'Port',
            'bytes': 'Total Bytes',
            'packets': 'Total Packets'
        })
        return result_df.sort_values('Total Bytes', ascending=False)
    
    def analyze_session_duration(self, top_n=20):
        """4. Top 20 session durations"""
        print(f"\n4. Analyzing top {top_n} session durations...")
        
        if self.df.empty or 'start' not in self.df.columns or 'end' not in self.df.columns:
            return pd.DataFrame({'Source IP': ['No data'], 'Destination IP': ['N/A'], 'Duration (seconds)': [0], 'Total Bytes': [0]})
        
        # Calculate duration in seconds (timestamps already converted in load_data)
        # Ensure both columns are datetime before calculation
        self.df['start'] = pd.to_datetime(self.df['start'], errors='coerce')
        self.df['end'] = pd.to_datetime(self.df['end'], errors='coerce')
        self.df['duration'] = (self.df['end'] - self.df['start']).dt.total_seconds()
        
        # Filter out invalid durations
        valid_durations = self.df[self.df['duration'] > 0].copy()
        
        if len(valid_durations) == 0:
            return pd.DataFrame({'Source IP': ['No valid durations'], 'Destination IP': ['N/A'], 'Duration (seconds)': [0], 'Total Bytes': [0]})
        
        # Group by connection and get max duration
        session_stats = valid_durations.groupby(['srcaddr', 'dstaddr']).agg({
            'duration': 'max',
            'bytes': 'sum',
            'packets': 'sum'
        }).reset_index()
        
        session_stats = session_stats.sort_values('duration', ascending=False).head(top_n)
        
        result_df = session_stats[['srcaddr', 'dstaddr', 'duration', 'bytes', 'packets']].copy()
        result_df = result_df.rename(columns={
            'srcaddr': 'Source IP',
            'dstaddr': 'Destination IP',
            'duration': 'Duration (seconds)',
            'bytes': 'Total Bytes',
            'packets': 'Total Packets'
        })
        return result_df
    
    def analyze_top_connections_by_bytes(self, top_n=20):
        """5. Top 20 connections by total bytes (srcIP → dstIP)"""
        print(f"\n5. Analyzing top {top_n} connections by bytes...")
        
        if self.df.empty or 'bytes' not in self.df.columns:
            return pd.DataFrame({'Source IP': ['No data'], 'Destination IP': ['N/A'], 'Total Bytes': [0], 'Total Packets': [0], 'Flow Count': [0]})
        
        # Group by source and destination IP
        connection_stats = self.df.groupby(['srcaddr', 'dstaddr']).agg({
            'bytes': 'sum',
            'packets': 'sum'
        }).reset_index()
        
        # Add flow count
        flow_counts = self.df.groupby(['srcaddr', 'dstaddr']).size().reset_index(name='Flow Count')
        connection_stats = connection_stats.merge(flow_counts, on=['srcaddr', 'dstaddr'])
        
        # Sort by total bytes
        connection_stats = connection_stats.sort_values('bytes', ascending=False).head(top_n)
        
        result_df = connection_stats[['srcaddr', 'dstaddr', 'bytes', 'packets', 'Flow Count']].copy()
        result_df = result_df.rename(columns={
            'srcaddr': 'Source IP',
            'dstaddr': 'Destination IP',
            'bytes': 'Total Bytes',
            'packets': 'Total Packets'
        })
        return result_df
    
    def save_analysis_to_excel(self, output_file):
        """Save analysis results to Excel file"""
        print(f"\n[INFO] Creating Excel file: {output_file}")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 1. Top 20 IPs by bytes
            top_ips = self.analyze_top_ips_by_bytes(20)
            top_ips.to_excel(writer, sheet_name='Access_IP_Statistics_(Top_20)', index=False)
            
            # 2. Top ports and protocols
            top_ports = self.analyze_top_ports_protocols(20)
            top_ports.to_excel(writer, sheet_name='Port_Protocol_Statistics', index=False)
            
            # 3. Nightshift remote access events
            nightshift_remote = self.analyze_nightshift_remote_access()
            nightshift_remote.to_excel(writer, sheet_name='Nightshift_Remote_Access_Events', index=False)
            
            # 4. Top session durations
            session_duration = self.analyze_session_duration(20)
            session_duration.to_excel(writer, sheet_name='Session_Duration_Statistics', index=False)
            
            # 5. Top connections by bytes
            top_connections = self.analyze_top_connections_by_bytes(20)
            top_connections.to_excel(writer, sheet_name='Connection_Statistics_(Top_20)', index=False)
        
        # Excel file formatting
        self._format_excel_sheets(output_file)
        
        print(f"[SUCCESS] Excel file created successfully: {output_file}")
        print(f"[INFO] Total sheets: 5 (5 analysis sheets)")
        
        return {
            'top_ips': top_ips,
            'top_ports': top_ports,
            'nightshift_remote': nightshift_remote,
            'session_duration': session_duration,
            'top_connections': top_connections
        }
    
    def _format_excel_sheets(self, output_file):
        """Format Excel sheets with styling"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = load_workbook(output_file)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Header styling
                header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                header_font = Font(bold=True)
                
                # Remove borders from header row
                no_border = Border()
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = no_border
                
                # Auto-adjust column widths
                self._adjust_column_widths(ws)
            
            wb.save(output_file)
            print("[SUCCESS] Excel formatting applied successfully")
            
        except Exception as e:
            print(f"[WARNING] Error during Excel formatting: {e}")
    
    def _adjust_column_widths(self, ws):
        """Adjust column widths for better readability"""
        try:
            from openpyxl.utils import get_column_letter
            
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            print(f"[WARNING] Error adjusting column widths: {e}")
    
    def generate_html_report(self, report_dir, analysis_data=None):
        """Generate HTML report"""
        print(f"\n[INFO] Generating HTML report...")
        
        if analysis_data is None:
            # Generate analysis data
            top_ips = self.analyze_top_ips_by_bytes(20)
            top_ports = self.analyze_top_ports_protocols(20)
            nightshift_remote = self.analyze_nightshift_remote_access()
            session_duration = self.analyze_session_duration(20)
            top_connections = self.analyze_top_connections_by_bytes(20)
        else:
            # Use provided analysis data
            top_ips = analysis_data['top_ips']
            top_ports = analysis_data['top_ports']
            nightshift_remote = analysis_data['nightshift_remote']
            session_duration = analysis_data['session_duration']
            top_connections = analysis_data['top_connections']
        
        # Generate HTML content
        html_content = self._create_html_content(
            top_ips, top_ports, nightshift_remote, session_duration, top_connections
        )
        
        # Save HTML file (report folder)
        os.makedirs(report_dir, exist_ok=True)
        
        html_filename = "vpc_flow_analysis_report.html"
        html_file = os.path.join(report_dir, html_filename)
        
        with open(html_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"[SUCCESS] HTML report created: {html_file}")
        return html_file
    
    def _create_html_content(self, top_ips, top_ports, nightshift_remote, session_duration, top_connections):
        """Create HTML content for the report"""
        
        html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Summary Analysis Report (VPC Flow Log)</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
            color: #333;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 10px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}
        .header h1 {{
            margin: 0;
            font-size: 2.5em;
            font-weight: 300;
        }}
        .header p {{
            margin: 10px 0 0 0;
            opacity: 0.9;
            font-size: 1.1em;
        }}
        .section {{
            padding: 30px;
            border-bottom: 1px solid #eee;
        }}
        .section:last-child {{
            border-bottom: none;
        }}
        .section h2 {{
            color: #2c3e50;
            margin: 0 0 20px 0;
            font-size: 1.8em;
            border-left: 4px solid #3498db;
            padding-left: 15px;
        }}
        .chart-container {{
            margin: 20px 0;
            height: 400px;
            position: relative;
        }}
        .chart-wrapper {{
            width: 100%;
            height: 400px;
            margin: 20px 0;
            background: #fafafa;
            border-radius: 8px;
            padding: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        th {{
            background: #34495e;
            color: white;
            padding: 15px;
            text-align: left;
            font-weight: 600;
        }}
        td {{
            padding: 12px 15px;
            border-bottom: 1px solid #eee;
        }}
        tr:nth-child(even) {{
            background: #f8f9fa;
        }}
        tr:hover {{
            background: #e3f2fd;
        }}
        .no-data {{
            text-align: center;
            color: #7f8c8d;
            font-style: italic;
            padding: 40px;
        }}
        .footer {{
            background: #2c3e50;
            color: white;
            text-align: center;
            padding: 20px;
            font-size: 0.9em;
        }}
    </style>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Summary Analysis Report (VPC Flow Log)</h1>
            <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
        
        {self._create_top_ips_section(top_ips)}
        {self._create_top_ports_section(top_ports)}
        {self._create_nightshift_remote_section(nightshift_remote)}
        {self._create_session_duration_section(session_duration)}
        {self._create_top_connections_section(top_connections)}
        
        <div class="footer">
            <p>Summary Analysis Report (VPC Flow Log) - Generated by PLAINBIT Co., LTD.</p>
        </div>
    </div>
</body>
</html>
        """
        return html
    
    def _create_top_ips_section(self, top_ips):
        """Create top IPs section"""
        if top_ips.empty or 'No data' in top_ips['IP'].values:
            return f"""
        <div class="section">
            <h2>1. Access IP Statistics (Top 20)</h2>
            <div class="no-data">No IP data available</div>
        </div>
            """
        
        # Chart data preparation
        chart_data = top_ips.head(10)
        chart_labels = [str(ip) for ip in chart_data['IP']]
        chart_values = chart_data['Total Bytes'].tolist()
        
        table_rows = ""
        for _, row in top_ips.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{row['IP']}</td>
                <td>{row['Total Bytes']:,}</td>
                <td>{row['Type']}</td>
                <td>{row['Total Packets']:,}</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>1. Access IP Statistics (Top 20)</h2>
            <div class="chart-wrapper">
                <canvas id="topIpsChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>IP Address</th>
                        <th>Total Bytes</th>
                        <th>Type</th>
                        <th>Total Packets</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const topIpsCtx = document.getElementById('topIpsChart').getContext('2d');
                new Chart(topIpsCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)',
                                'rgba(39, 174, 96, 0.8)',
                                'rgba(41, 128, 185, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Top 10 IPs by Total Bytes'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def _create_top_ports_section(self, top_ports):
        """Create top ports section"""
        if top_ports.empty:
            return f"""
        <div class="section">
            <h2>2. Port Protocol Statistics (Top 20)</h2>
            <div class="no-data">No port data available</div>
        </div>
            """
        
        # Chart data preparation
        chart_data = top_ports.head(10)
        chart_labels = [f"Port {int(port)}" for port in chart_data['Port']]
        chart_values = chart_data['Total Bytes'].tolist()
        
        table_rows = ""
        for _, row in top_ports.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{int(row['Port'])}</td>
                <td>{row['Protocol']}</td>
                <td>{row['Flow Count']:,}</td>
                <td>{row['Total Bytes']:,}</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>2. Port Protocol Statistics (Top 20)</h2>
            <div class="chart-wrapper">
                <canvas id="topPortsChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>Port</th>
                        <th>Protocol</th>
                        <th>Flow Count</th>
                        <th>Total Bytes</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const topPortsCtx = document.getElementById('topPortsChart').getContext('2d');
                new Chart(topPortsCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)',
                                'rgba(39, 174, 96, 0.8)',
                                'rgba(41, 128, 185, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Top 10 Ports by Total Bytes'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def _create_nightshift_remote_section(self, nightshift_remote):
        """Create nightshift remote access section"""
        if nightshift_remote.empty or 'No nightshift events' in nightshift_remote['Source IP'].values:
            return f"""
        <div class="section">
            <h2>3. Nightshift Remote Access Events (22:00-06:00)</h2>
            <div class="no-data">No nightshift remote access events</div>
        </div>
            """
        
        # Chart data preparation
        chart_data = nightshift_remote.head(10)
        chart_labels = [f"{row['Source IP']} → {row['Destination IP']}" for _, row in chart_data.iterrows()]
        chart_values = chart_data['Total Bytes'].tolist()
        
        table_rows = ""
        for _, row in nightshift_remote.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{row['Source IP']}</td>
                <td>{row['Destination IP']}</td>
                <td>{int(row['Port'])}</td>
                <td>{row['Protocol']}</td>
                <td>{row['Service']}</td>
                <td>{row['Total Bytes']:,}</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>3. Nightshift Remote Access Events (22:00-06:00)</h2>
            <div class="chart-wrapper">
                <canvas id="nightshiftRemoteChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>Source IP</th>
                        <th>Destination IP</th>
                        <th>Port</th>
                        <th>Protocol</th>
                        <th>Service</th>
                        <th>Total Bytes</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const nightshiftRemoteCtx = document.getElementById('nightshiftRemoteChart').getContext('2d');
                new Chart(nightshiftRemoteCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)',
                                'rgba(39, 174, 96, 0.8)',
                                'rgba(41, 128, 185, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Top 10 Nightshift Remote Access Events'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def _create_session_duration_section(self, session_duration):
        """Create session duration section"""
        if session_duration.empty or 'No valid durations' in session_duration['Source IP'].values:
            return f"""
        <div class="section">
            <h2>4. Session Duration Statistics (Top 20)</h2>
            <div class="no-data">No session duration data available</div>
        </div>
            """
        
        # Chart data preparation
        chart_data = session_duration.head(10)
        chart_labels = [f"{row['Source IP']} → {row['Destination IP']}" for _, row in chart_data.iterrows()]
        chart_values = chart_data['Duration (seconds)'].tolist()
        
        table_rows = ""
        for _, row in session_duration.head(20).iterrows():
            duration_hours = row['Duration (seconds)'] / 3600
            table_rows += f"""
            <tr>
                <td>{row['Source IP']}</td>
                <td>{row['Destination IP']}</td>
                <td>{duration_hours:.2f} hours</td>
                <td>{row['Total Bytes']:,}</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>4. Session Duration Statistics (Top 20)</h2>
            <div class="chart-wrapper">
                <canvas id="sessionDurationChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>Source IP</th>
                        <th>Destination IP</th>
                        <th>Duration</th>
                        <th>Total Bytes</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const sessionDurationCtx = document.getElementById('sessionDurationChart').getContext('2d');
                new Chart(sessionDurationCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)',
                                'rgba(39, 174, 96, 0.8)',
                                'rgba(41, 128, 185, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Top 10 Session Durations (seconds)'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def _create_top_connections_section(self, top_connections):
        """Create top connections section"""
        if top_connections.empty or 'No data' in top_connections['Source IP'].values:
            return f"""
        <div class="section">
            <h2>5. Connection Statistics (Top 20)</h2>
            <div class="no-data">No connection data available</div>
        </div>
            """
        
        # Chart data preparation
        chart_data = top_connections.head(10)
        chart_labels = [f"{row['Source IP']} → {row['Destination IP']}" for _, row in chart_data.iterrows()]
        chart_values = chart_data['Total Bytes'].tolist()
        
        table_rows = ""
        for _, row in top_connections.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{row['Source IP']}</td>
                <td>{row['Destination IP']}</td>
                <td>{row['Total Bytes']:,}</td>
                <td>{row['Total Packets']:,}</td>
                <td>{row['Flow Count']:,}</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>5. Connection Statistics (Top 20)</h2>
            <div class="chart-wrapper">
                <canvas id="topConnectionsChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>Source IP</th>
                        <th>Destination IP</th>
                        <th>Total Bytes</th>
                        <th>Total Packets</th>
                        <th>Flow Count</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const topConnectionsCtx = document.getElementById('topConnectionsChart').getContext('2d');
                new Chart(topConnectionsCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)',
                                'rgba(39, 174, 96, 0.8)',
                                'rgba(41, 128, 185, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Top 10 Connections by Total Bytes'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """


def find_vpc_flow_csv():
    """Find VPC Flow Log CSV file"""
    # Find latest VPC Flow Log CSV file in result/parse_log folder
    parse_log_dir = os.path.join('result', 'parse_log')
    if os.path.exists(parse_log_dir):
        csv_files = [f for f in os.listdir(parse_log_dir) if f.startswith('vpc_flow_log_') and f.endswith('.csv')]
        if csv_files:
            latest_csv = os.path.join(parse_log_dir, max(csv_files, key=lambda x: os.path.getctime(os.path.join(parse_log_dir, x))))
            return latest_csv
    
    # Find latest VPC Flow Log CSV file in output folder (legacy compatibility)
    output_dir = 'output'
    if os.path.exists(output_dir):
        csv_files = [f for f in os.listdir(output_dir) if f.startswith('vpc_flow_log_') and f.endswith('.csv')]
        if csv_files:
            latest_csv = os.path.join(output_dir, max(csv_files, key=lambda x: os.path.getctime(os.path.join(output_dir, x))))
            return latest_csv
    
    # Search in current folder
    csv_files = [f for f in os.listdir('.') if f.startswith('vpc_flow_log_') and f.endswith('.csv')]
    if csv_files:
        return max(csv_files, key=os.path.getctime)
    
    return None


def run_vpc_flow_parser():
    """Run VPC Flow Log parser"""
    print("🔄 Starting VPC Flow Log parsing...")
    
    try:
        # VPC Flow Log parser import and execution
        from vpc_flow_log_parser import VPCFlowLogParser
        
        # Default path setting
        input_path = r"C:\Users\kelly.jang\Desktop\새 폴더\AWS_LOG\VPC Flow Log (S3)"
        
        # Directory structure creation
        result_dir = "result"
        parse_log_dir = os.path.join(result_dir, "parse_log")
        
        # Output directory creation
        os.makedirs(parse_log_dir, exist_ok=True)
        
        # Parser execution
        parser = VPCFlowLogParser()
        events = parser.parse_directory(input_path)
        
        if events:
            # Output filename generation
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(parse_log_dir, f"vpc_flow_log_{timestamp}.csv")
            
            # CSV save
            parser.save_to_csv(events, output_file)
            print(f"[SUCCESS] VPC Flow Log parsing completed: {output_file}")
            return output_file
        else:
            print("[ERROR] No events parsed.")
            return None
            
    except Exception as e:
        print(f"[ERROR] Error during VPC Flow Log parsing: {e}")
        return None


def main():
    """Main function"""
    print("[INFO] VPC Flow Log Analyzer")
    print("=" * 50)
    
    # Directory structure creation
    result_dir = "result"
    parse_log_dir = os.path.join(result_dir, "parse_log")
    analysis_dir = os.path.join(result_dir, "analysis")
    report_dir = os.path.join(result_dir, "report")
    
    # Directory creation
    os.makedirs(parse_log_dir, exist_ok=True)
    os.makedirs(analysis_dir, exist_ok=True)
    os.makedirs(report_dir, exist_ok=True)
    
    # 1. VPC Flow Log CSV file search
    latest_csv = find_vpc_flow_csv()
    
    if not latest_csv:
        print("📁 VPC Flow Log CSV file not found.")
        print("🔄 Starting VPC Flow Log parsing...")
        
        # VPC Flow Log parser execution
        latest_csv = run_vpc_flow_parser()
        
        if not latest_csv:
            print("❌ VPC Flow Log parsing failed.")
            return
    else:
        print(f"📁 Analysis file: {latest_csv}")
    
    # 2. Analyzer creation and execution
    print(f"\n[INFO] Starting VPC Flow Log analysis...")
    analyzer = VPCFlowLogAnalyzer(latest_csv)
    
    # 3. Output filename generation
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(analysis_dir, f"vpc_flow_analysis_{timestamp}.xlsx")
    
    # 4. Excel file creation and analysis data collection
    analysis_data = analyzer.save_analysis_to_excel(output_file)
    
    # 5. HTML report generation (reuse analysis data)
    html_file = analyzer.generate_html_report(output_file, analysis_data)
    
    print(f"\n[COMPLETE] VPC Flow Log analysis completed!")
    print(f"[INFO] Excel file: {output_file}")
    print(f"[INFO] HTML report: {html_file}")
    print(f"[INFO] Check the detailed analysis results with 5 sheets and HTML report!")
    
    return output_file, html_file


def run_analysis(csv_file: str, analysis_dir: str, report_dir: str) -> tuple:
    """Run complete VPC Flow Log analysis and return file paths"""
    try:
        # Create directories
        os.makedirs(analysis_dir, exist_ok=True)
        os.makedirs(report_dir, exist_ok=True)
        
        # Create analyzer instance with CSV file path
        analyzer = VPCFlowLogAnalyzer(csv_file)
        
        # Generate output filename for Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"vpc_flow_analysis_{timestamp}.xlsx"
        excel_file = os.path.join(analysis_dir, excel_filename)
        
        # Run analysis and save to Excel
        analysis_data = analyzer.save_analysis_to_excel(excel_file)
        
        # Generate HTML report
        html_file = analyzer.generate_html_report(report_dir, analysis_data)
        
        return excel_file, html_file
        
    except Exception as e:
        print(f"[ERROR] VPC Flow Log analysis failed: {str(e)}")
        raise


if __name__ == "__main__":
    main()
