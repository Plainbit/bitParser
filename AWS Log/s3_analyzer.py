#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import json
import pandas as pd
from datetime import datetime
import argparse
import re

# Windows Korean encoding fix
import codecs
import locale

# Set encoding for Windows
try:
    os.environ['PYTHONIOENCODING'] = 'utf-8'
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    else:
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.detach())
except:
    os.environ['PYTHONIOENCODING'] = 'utf-8'

try:
    locale.setlocale(locale.LC_ALL, 'Korean_Korea.949')
except:
    try:
        locale.setlocale(locale.LC_ALL, 'ko_KR.UTF-8')
    except:
        pass


class S3AccessLogAnalyzer:
    def __init__(self, csv_file):
        """Initialize S3 Access Log Analyzer"""
        print(f"Loading S3 Access Log data from: {csv_file}")
        
        # Load CSV data
        self.df = pd.read_csv(csv_file)
        print(f"Loaded data: {len(self.df)} access log records")
        
        # Load MITRE ATT&CK configuration
        self.mitre_config = self._load_mitre_config()
        
        # Convert timestamp columns to datetime
        if 'request_datetime_(UTC+0)' in self.df.columns:
            # Handle new column format: YYYY-MM-DD hh:mm:ss
            self.df['Timestamp'] = pd.to_datetime(self.df['request_datetime_(UTC+0)'], errors='coerce')
            # Remove timezone info for Excel compatibility
            self.df['Timestamp'] = self.df['Timestamp'].dt.tz_localize(None)
        elif 'request_datetime' in self.df.columns:
            # Fallback for old format
            self.df['request_datetime'] = self.df['request_datetime'].apply(self._convert_s3_timestamp)
            self.df['Timestamp'] = pd.to_datetime(self.df['request_datetime'], errors='coerce')
            # Remove timezone info for Excel compatibility
            self.df['Timestamp'] = self.df['Timestamp'].dt.tz_localize(None)
        
        if 'datetime_local' in self.df.columns:
            # Convert S3 timestamp format [01/Oct/2025:11:55:11 +0000] to YYYY-MM-DD hh:mm:ss
            self.df['datetime_local'] = self.df['datetime_local'].apply(self._convert_s3_timestamp)
            # Remove timezone info for Excel compatibility
            self.df['datetime_local'] = pd.to_datetime(self.df['datetime_local'], errors='coerce')
            self.df['datetime_local'] = self.df['datetime_local'].dt.tz_localize(None)
    
    def _load_mitre_config(self):
        """Load MITRE ATT&CK configuration"""
        try:
            with open('mitre_config.json', 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            print("Warning: mitre_config.json not found. MITRE ATT&CK analysis will be skipped.")
            return {}
    
    def analyze_requester_ips(self, top_n=20):
        """1. Analyze top requesters and source IPs (separated ARN and IP)"""
        print(f"\n1. Analyzing top {top_n} requesters and source IPs...")
        
        if self.df.empty:
            return pd.DataFrame({'Requester ARN': ['No data'], 'Source IP': ['N/A'], 'Request Count': [0]})
        
        # Extract ARN and IP from requester field
        requester_info = self.df['requester'].apply(self._extract_requester_info)
        self.df['requester_arn'] = [info[0] for info in requester_info]
        self.df['requester_ip'] = [info[1] for info in requester_info]
        
        # Use remote_ip as primary IP source, fallback to requester_ip
        self.df['final_ip'] = self.df['remote_ip'].fillna(self.df['requester_ip'])
        
        # Group by requester ARN and final IP
        requester_stats = self.df.groupby(['requester_arn', 'final_ip']).size().reset_index(name='Request Count')
        requester_stats = requester_stats.sort_values('Request Count', ascending=False).head(top_n)
        
        # Filter out N/A values
        requester_stats = requester_stats[
            (requester_stats['requester_arn'] != 'N/A') & 
            (requester_stats['final_ip'] != 'N/A')
        ]
        
        # Rename columns
        requester_stats.columns = ['Requester ARN', 'Source IP', 'Request Count']
        
        return requester_stats
    
    def analyze_operations_by_bucket_prefix(self, top_n=20):
        """2. Analyze PUT/GET/DELETE/COPY operations by bucket and prefix"""
        print(f"\n2. Analyzing operations by bucket and prefix...")
        
        if self.df.empty:
            return pd.DataFrame({'Operation': ['No data'], 'Bucket': ['N/A'], 'Prefix': ['N/A'], 'Count': [0]})
        
        # Filter for main operations (S3 REST API format)
        operations = ['REST.PUT.OBJECT', 'REST.GET.OBJECT', 'REST.DELETE.OBJECT', 'REST.COPY.OBJECT']
        operation_data = self.df[self.df['operation'].isin(operations)].copy()
        
        if operation_data.empty:
            return pd.DataFrame({'Operation': ['No operations'], 'Bucket': ['N/A'], 'Prefix': ['N/A'], 'Count': [0]})
        
        # Group by operation, bucket, and prefix
        operation_stats = operation_data.groupby(['operation', 'bucket', 'key']).size().reset_index(name='Count')
        operation_stats = operation_stats.sort_values('Count', ascending=False).head(top_n)
        
        # Clean up prefix (remove leading slash)
        operation_stats['key'] = operation_stats['key'].str.lstrip('/')
        operation_stats['key'] = operation_stats['key'].fillna('(root)')
        
        # Rename columns
        operation_stats.columns = ['Operation', 'Bucket', 'Prefix', 'Count']
        
        return operation_stats
    
    def analyze_user_agents(self, top_n=20):
        """3. Analyze UserAgent statistics (detailed classification)"""
        print(f"\n3. Analyzing UserAgent statistics...")
        
        if self.df.empty:
            return pd.DataFrame({'UserAgent Type': ['No data'], 'Usage Count': [0], 'Usage Statistics (%)': [0]})
        
        # Classify all UserAgents first
        self.df['user_agent_type'] = self.df['user_agent'].apply(self._classify_user_agent)
        
        # Count by UserAgent Type (not individual UserAgent strings)
        ua_type_counts = self.df['user_agent_type'].value_counts().head(top_n)
        total_requests = len(self.df)
        
        ua_stats = []
        for ua_type, count in ua_type_counts.items():
            percentage = (count / total_requests) * 100
            ua_stats.append({
                'UserAgent Type': ua_type,
                'Usage Count': count,
                'Usage Statistics (%)': round(percentage, 2)
            })
        
        return pd.DataFrame(ua_stats)
    
    def _classify_user_agent(self, user_agent):
        """Classify UserAgent into categories"""
        if pd.isna(user_agent) or user_agent == '':
            return 'Unknown'
        
        ua_lower = user_agent.lower()
        
        # AWS CLI
        if 'aws-cli' in ua_lower or 'aws-sdk' in ua_lower:
            return 'AWS CLI/SDK'
        
        # Browsers
        if any(browser in ua_lower for browser in ['chrome', 'firefox', 'safari', 'edge', 'opera']):
            return 'Browser'
        
        # Mobile
        if any(mobile in ua_lower for mobile in ['mobile', 'android', 'iphone', 'ipad']):
            return 'Mobile'
        
        # Cloud services
        if any(cloud in ua_lower for cloud in ['cloudfront', 's3', 'lambda', 'ec2']):
            return 'Cloud Service'
        
        # Tools
        if any(tool in ua_lower for tool in ['curl', 'wget', 'postman', 'insomnia']):
            return 'Tool'
        
        # Applications
        if any(app in ua_lower for app in ['python', 'java', 'node', 'go', 'php']):
            return 'Application'
        
        return 'Other'
    
    def _convert_s3_timestamp(self, timestamp_str):
        """Convert S3 timestamp to YYYY-MM-DD hh:mm:ss format"""
        try:
            import re
            # Parse S3 Access Log timestamp format using regex: [01/Oct/2025:11:55:11 +0000]
            pattern = r'\[(\d{2})/(\w{3})/(\d{4}):(\d{2}):(\d{2}):(\d{2})\s+([+-]\d{4})\]'
            match = re.match(pattern, str(timestamp_str))
            
            if match:
                day, month, year, hour, minute, second, tz_offset = match.groups()
                
                # Convert month name to number
                month_map = {
                    'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
                    'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
                    'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
                }
                month_num = month_map.get(month, '01')
                
                # Return in YYYY-MM-DD hh:mm:ss format
                return f"{year}-{month_num}-{day} {hour}:{minute}:{second}"
            else:
                return timestamp_str
        except Exception as e:
            print(f"S3 timestamp conversion error ({timestamp_str}): {e}")
            return timestamp_str
    
    def _extract_requester_info(self, requester):
        """Extract ARN and IP from requester field using regex"""
        if pd.isna(requester) or requester == '' or requester == '-':
            return 'Unknown', 'Unknown'
        
        # ARN pattern: arn:aws:iam::account:role/role-name or arn:aws:sts::account:assumed-role/role-name/session-name
        arn_pattern = r'arn:aws:(iam|sts)::\d+:([^/]+)/(.+)'
        arn_match = re.search(arn_pattern, str(requester))
        
        # IP pattern: IPv4 or IPv6
        ip_pattern = r'\b(?:\d{1,3}\.){3}\d{1,3}\b|\b(?:[0-9a-fA-F]{1,4}:){7}[0-9a-fA-F]{1,4}\b'
        ip_match = re.search(ip_pattern, str(requester))
        
        if arn_match:
            service = arn_match.group(1)
            role_type = arn_match.group(2)
            role_name = arn_match.group(3)
            arn_display = f"{service}:{role_type}/{role_name}"
            return arn_display, 'N/A'
        elif ip_match:
            return 'N/A', ip_match.group(0)
        else:
            # If it's not ARN or IP, treat as ARN-like identifier
            return str(requester), 'N/A'
    
    def analyze_mitre_attack_tactics(self):
        """4. Analyze MITRE ATT&CK tactics (CloudTrail style)"""
        print(f"\n4. Analyzing MITRE ATT&CK tactics...")
        
        if not self.mitre_config or self.df.empty:
            return pd.DataFrame({'MITRE ATT&CK Tactics': ['No data'], 'Operation': ['N/A'], 'Operation Hit Count': [0], 'Operation Statistics (%)': [0]})
        
        # Create operation to tactic mapping (S3 Access Log specific)
        operation_to_tactic = {}
        s3_config = self.mitre_config.get('s3_access_log', {})
        for tactic, operations in s3_config.items():
            for operation in operations:
                operation_to_tactic[operation] = tactic
        
        # Count operations by tactic
        tactic_counts = {}
        total_operations = len(self.df)
        
        for _, row in self.df.iterrows():
            operation = row['operation']
            tactic = operation_to_tactic.get(operation, 'No Tactic Mapped')
            
            if tactic not in tactic_counts:
                tactic_counts[tactic] = {}
            
            if operation not in tactic_counts[tactic]:
                tactic_counts[tactic][operation] = 0
            tactic_counts[tactic][operation] += 1
        
        # Create results DataFrame
        results = []
        for tactic, operations in tactic_counts.items():
            if tactic == 'No Tactic Mapped':
                continue  # Skip unmapped tactics
            for operation, count in operations.items():
                percentage = (count / total_operations) * 100
                results.append({
                    'MITRE ATT&CK Tactics': tactic,
                    'Operation': operation,
                    'Operation Hit Count': count,
                    'Operation Statistics (%)': round(percentage, 2)
                })
        
        # Sort by hit count descending
        results_df = pd.DataFrame(results)
        results_df = results_df.sort_values('Operation Hit Count', ascending=False)
        
        return results_df
    
    def get_mitre_attack_events(self):
        """5. Get detailed MITRE ATT&CK events"""
        print(f"\n5. Getting detailed MITRE ATT&CK events...")
        
        if not self.mitre_config or self.df.empty:
            return pd.DataFrame({'MITRE ATT&CK': ['No data'], 'Operation': ['N/A'], 'Bucket': ['N/A'], 'Key': ['N/A'], 'Source IP': ['N/A'], 'Timestamp': ['N/A']})
        
        # Create operation to tactic mapping (S3 Access Log specific)
        operation_to_tactic = {}
        s3_config = self.mitre_config.get('s3_access_log', {})
        for tactic, operations in s3_config.items():
            for operation in operations:
                operation_to_tactic[operation] = tactic
        
        # Map operations to tactics
        self.df['mitre_tactic'] = self.df['operation'].map(operation_to_tactic)
        
        # Filter for mapped events
        mapped_events = self.df[self.df['mitre_tactic'].notna()].copy()
        
        if mapped_events.empty:
            return pd.DataFrame({'MITRE ATT&CK': ['No mapped events'], 'Operation': ['N/A'], 'Bucket': ['N/A'], 'Key': ['N/A'], 'Source IP': ['N/A'], 'Timestamp': ['N/A']})
        
        # Select all S3 log columns except mitre_tactic (which will be renamed)
        all_columns = [col for col in mapped_events.columns if col != 'mitre_tactic']
        result_df = mapped_events[all_columns].copy()
        
        # Add MITRE ATT&CK column at the beginning
        result_df.insert(0, 'MITRE ATT&CK', mapped_events['mitre_tactic'])
        
        # Rename some key columns for better readability
        column_mapping = {
            'operation': 'Operation',
            'bucket': 'Bucket', 
            'key': 'Key',
            'remote_ip': 'Source IP',
            'request_datetime_(UTC+0)': 'Timestamp (UTC)',
            'timezone': 'Timezone',
            'datetime_local': 'Timestamp (Local)',
            'requester': 'Requester',
            'user_agent': 'User Agent',
            'request_uri': 'Request URI',
            'http_status': 'HTTP Status',
            'error_code': 'Error Code',
            'bytes_sent': 'Bytes Sent',
            'object_size': 'Object Size',
            'total_time': 'Total Time',
            'turn_around_time': 'Turn Around Time',
            'referer': 'Referer'
        }
        
        # Only rename columns that exist
        existing_mapping = {k: v for k, v in column_mapping.items() if k in result_df.columns}
        result_df = result_df.rename(columns=existing_mapping)
        
        # Convert timestamps to string to avoid Excel issues
        if 'Timestamp (UTC)' in result_df.columns:
            result_df['Timestamp (UTC)'] = result_df['Timestamp (UTC)'].astype(str)
        if 'Timestamp (Local)' in result_df.columns:
            result_df['Timestamp (Local)'] = result_df['Timestamp (Local)'].astype(str)
        
        return result_df
    
    def save_analysis_to_excel(self, output_file):
        """Save analysis results to Excel file"""
        print(f"\n[INFO] Creating Excel file: {output_file}")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 1. Top requesters and source IPs
            requester_ips = self.analyze_requester_ips(20)
            requester_ips.to_excel(writer, sheet_name='Access_IP_Statistics_(Top_20)', index=False)
            
            # 2. Operations by bucket and prefix
            operations = self.analyze_operations_by_bucket_prefix(20)
            operations.to_excel(writer, sheet_name='Operation_Statistics', index=False)
            
            # 3. UserAgent statistics
            user_agents = self.analyze_user_agents(20)
            user_agents.to_excel(writer, sheet_name='UserAgent_Statistics', index=False)
            
            # 4. MITRE ATT&CK tactics
            mitre_tactics = self.analyze_mitre_attack_tactics()
            mitre_tactics.to_excel(writer, sheet_name='Mitre_ATTACK_Tactics', index=False)
            
            # 5. MITRE ATT&CK events detail
            mitre_events = self.get_mitre_attack_events()
            mitre_events.to_excel(writer, sheet_name='Mitre_ATTACK_Events', index=False)
        
        # Excel file formatting
        self._format_excel_sheets(output_file)
        
        print(f"[SUCCESS] Excel file created successfully: {output_file}")
        print(f"[INFO] Total sheets: 5 (5 analysis sheets)")
        
        return {
            'requester_ips': requester_ips,
            'operations': operations,
            'user_agents': user_agents,
            'mitre_tactics': mitre_tactics,
            'mitre_events': mitre_events
        }
    
    def _format_excel_sheets(self, output_file):
        """Format Excel sheets with styling"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            wb = load_workbook(output_file)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Header formatting
                header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                header_font = Font(bold=True)
                
                # Apply header formatting to first row
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # Remove borders from header
                    cell.border = Border()
                
                # Auto-adjust column widths
                self._adjust_column_widths(ws)
            
            wb.save(output_file)
            print("[SUCCESS] Excel formatting applied successfully")
            
        except Exception as e:
            print(f"[WARNING] Excel formatting failed: {e}")
    
    def _adjust_column_widths(self, worksheet):
        """Adjust column widths for better readability"""
        try:
            from openpyxl.utils import get_column_letter
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            print(f"[WARNING] Column width adjustment failed: {e}")
    
    def generate_html_report(self, analysis_data, output_file):
        """Generate HTML report"""
        print(f"\n[INFO] Generating HTML report...")
        
        html_content = self._create_html_content(analysis_data)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"[SUCCESS] HTML report created: {output_file}")
    
    def _create_html_content(self, analysis_data):
        """Create HTML content"""
        html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Summary Analysis Report (S3 Access Log)</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
            color: #333;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 10px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}
        .header h1 {{
            margin: 0;
            font-size: 2.5em;
            font-weight: 300;
        }}
        .header p {{
            margin: 10px 0 0 0;
            opacity: 0.9;
            font-size: 1.1em;
        }}
        .section {{
            padding: 30px;
            border-bottom: 1px solid #eee;
        }}
        .section:last-child {{
            border-bottom: none;
        }}
        .section h2 {{
            color: #2c3e50;
            margin: 0 0 20px 0;
            font-size: 1.8em;
            border-left: 4px solid #3498db;
            padding-left: 15px;
        }}
        .chart-wrapper {{
            width: 100%;
            height: 400px;
            margin: 20px 0;
            background: #fafafa;
            border-radius: 8px;
            padding: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        th {{
            background: #34495e;
            color: white;
            padding: 15px;
            text-align: left;
            font-weight: 600;
        }}
        td {{
            padding: 12px 15px;
            border-bottom: 1px solid #eee;
        }}
        tr:nth-child(even) {{
            background: #f8f9fa;
        }}
        tr:hover {{
            background: #e3f2fd;
        }}
        .no-data {{
            text-align: center;
            color: #7f8c8d;
            font-style: italic;
            padding: 40px;
        }}
        .footer {{
            background: #2c3e50;
            color: white;
            text-align: center;
            padding: 20px;
            font-size: 0.9em;
        }}
    </style>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Summary Analysis Report (S3 Access Log)</h1>
            <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
        
        {self._create_requester_ips_section(analysis_data['requester_ips'])}
        {self._create_operations_section(analysis_data['operations'])}
        {self._create_user_agents_section(analysis_data['user_agents'])}
        {self._create_mitre_tactics_section(analysis_data['mitre_tactics'])}
        
        <div class="footer">
            <p>Summary Analysis Report (S3 Access Log) - Generated by PLAINBIT Co., LTD.</p>
        </div>
    </div>
</body>
</html>
        """
        return html
    
    def _create_requester_ips_section(self, requester_ips):
        """Create requester IPs section"""
        if requester_ips.empty or 'No data' in requester_ips['Requester ARN'].values:
            return f"""
        <div class="section">
            <h2>1. Requester IP Statistics (Top 20)</h2>
            <div class="no-data">No requester IP data available</div>
        </div>
            """
        
        # Chart data preparation
        chart_data = requester_ips.head(10)
        chart_labels = [f"{row['Requester ARN']} ({row['Source IP']})" for _, row in chart_data.iterrows()]
        chart_values = chart_data['Request Count'].tolist()
        
        table_rows = ""
        for _, row in requester_ips.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{row['Requester ARN']}</td>
                <td>{row['Source IP']}</td>
                <td>{row['Request Count']:,}</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>1. Requester IP Statistics (Top 20)</h2>
            <div class="chart-wrapper">
                <canvas id="requesterIpsChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>Requester</th>
                        <th>Source IP</th>
                        <th>Request Count</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const requesterIpsCtx = document.getElementById('requesterIpsChart').getContext('2d');
                new Chart(requesterIpsCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)',
                                'rgba(39, 174, 96, 0.8)',
                                'rgba(41, 128, 185, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Top 10 Requesters by Request Count'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def _create_operations_section(self, operations):
        """Create operations section"""
        if operations.empty or 'No data' in operations['Operation'].values:
            return f"""
        <div class="section">
            <h2>2. Operation Statistics (Top 20)</h2>
            <div class="no-data">No operation data available</div>
        </div>
            """
        
        # Chart data preparation
        chart_data = operations.head(10)
        chart_labels = [f"{row['Operation']} - {row['Bucket']}" for _, row in chart_data.iterrows()]
        chart_values = chart_data['Count'].tolist()
        
        table_rows = ""
        for _, row in operations.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{row['Operation']}</td>
                <td>{row['Bucket']}</td>
                <td>{row['Prefix']}</td>
                <td>{row['Count']:,}</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>2. Operation Statistics (Top 20)</h2>
            <div class="chart-wrapper">
                <canvas id="operationsChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>Operation</th>
                        <th>Bucket</th>
                        <th>Prefix</th>
                        <th>Count</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const operationsCtx = document.getElementById('operationsChart').getContext('2d');
                new Chart(operationsCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)',
                                'rgba(39, 174, 96, 0.8)',
                                'rgba(41, 128, 185, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Top 10 Operations by Count'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def _create_user_agents_section(self, user_agents):
        """Create user agents section"""
        if user_agents.empty or 'No data' in user_agents['UserAgent Type'].values:
            return f"""
        <div class="section">
            <h2>3. UserAgent Statistics</h2>
            <div class="no-data">No UserAgent data available</div>
        </div>
            """
        
        # Chart data preparation
        chart_data = user_agents.head(10)
        chart_labels = [ua[:50] + '...' if len(ua) > 50 else ua for ua in chart_data['UserAgent Type']]
        chart_values = chart_data['Usage Count'].tolist()
        
        table_rows = ""
        for _, row in user_agents.head(20).iterrows():
            ua_display = row['UserAgent Type'][:100] + '...' if len(row['UserAgent Type']) > 100 else row['UserAgent Type']
            table_rows += f"""
            <tr>
                <td title="{row['UserAgent Type']}">{ua_display}</td>
                <td>{row['Usage Count']:,}</td>
                <td>{row['Usage Statistics (%)']}%</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>3. UserAgent Statistics</h2>
            <div class="chart-wrapper">
                <canvas id="userAgentsChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>UserAgent Type</th>
                        <th>Usage Count</th>
                        <th>Usage Statistics (%)</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const userAgentsCtx = document.getElementById('userAgentsChart').getContext('2d');
                new Chart(userAgentsCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)',
                                'rgba(39, 174, 96, 0.8)',
                                'rgba(41, 128, 185, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Top 10 UserAgents by Count'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def _create_mitre_tactics_section(self, mitre_tactics):
        """Create MITRE ATT&CK tactics section"""
        if mitre_tactics.empty or 'No data' in mitre_tactics['MITRE ATT&CK Tactics'].values:
            return f"""
        <div class="section">
            <h2>4. MITRE ATT&CK Tactics Statistics</h2>
            <div class="no-data">No MITRE ATT&CK tactics data available</div>
        </div>
            """
        
        # Chart data preparation
        chart_data = mitre_tactics.head(10)
        chart_labels = chart_data['MITRE ATT&CK Tactics'].tolist()
        chart_values = chart_data['Operation Hit Count'].tolist()
        
        table_rows = ""
        for _, row in mitre_tactics.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{row['MITRE ATT&CK Tactics']}</td>
                <td>{row['Operation']}</td>
                <td>{row['Operation Hit Count']:,}</td>
                <td>{row['Operation Statistics (%)']}%</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>4. MITRE ATT&CK Tactics Statistics</h2>
            <div class="chart-wrapper">
                <canvas id="mitreTacticsChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>MITRE ATT&CK Tactics</th>
                        <th>Operation</th>
                        <th>Operation Hit Count</th>
                        <th>Operation Statistics (%)</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const mitreTacticsCtx = document.getElementById('mitreTacticsChart').getContext('2d');
                new Chart(mitreTacticsCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)',
                                'rgba(39, 174, 96, 0.8)',
                                'rgba(41, 128, 185, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Top 10 MITRE ATT&CK Tactics by Hit Count'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def _create_mitre_events_section(self, mitre_events):
        """Create MITRE ATT&CK events section"""
        if mitre_events.empty or 'No data' in mitre_events['MITRE ATT&CK'].values:
            return f"""
        <div class="section">
            <h2>5. MITRE ATT&CK Event Details</h2>
            <div class="no-data">No MITRE ATT&CK events available</div>
        </div>
            """
        
        table_rows = ""
        for _, row in mitre_events.head(20).iterrows():
            key_display = row['Key'][:50] + '...' if len(str(row['Key'])) > 50 else row['Key']
            table_rows += f"""
            <tr>
                <td>{row['MITRE ATT&CK']}</td>
                <td>{row['Operation']}</td>
                <td>{row['Bucket']}</td>
                <td title="{row['Key']}">{key_display}</td>
                <td>{row['Source IP']}</td>
                <td>{row.get('Timestamp (UTC)', 'N/A')}</td>
                <td>{row.get('Timezone', 'N/A')}</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>5. MITRE ATT&CK Event Details</h2>
            <table>
                <thead>
                    <tr>
                        <th>MITRE ATT&CK</th>
                        <th>Operation</th>
                        <th>Bucket</th>
                        <th>Key</th>
                        <th>Source IP</th>
                        <th>Timestamp (UTC)</th>
                        <th>Timezone</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
        </div>
        """


def find_s3_access_csv():
    """Find S3 Access Log CSV file"""
    # Find latest S3 Access Log CSV file in result/parse_log folder
    parse_log_dir = os.path.join('result', 'parse_log')
    if os.path.exists(parse_log_dir):
        csv_files = [f for f in os.listdir(parse_log_dir) if f.startswith('s3_access_log_') and f.endswith('.csv')]
        if csv_files:
            latest_csv = os.path.join(parse_log_dir, max(csv_files, key=lambda x: os.path.getctime(os.path.join(parse_log_dir, x))))
            return latest_csv
    
    # Find latest S3 Access Log CSV file in output folder (legacy compatibility)
    output_dir = 'output'
    if os.path.exists(output_dir):
        csv_files = [f for f in os.listdir(output_dir) if f.startswith('s3_access_log_') and f.endswith('.csv')]
        if csv_files:
            latest_csv = os.path.join(output_dir, max(csv_files, key=lambda x: os.path.getctime(os.path.join(output_dir, x))))
            return latest_csv
    
    # Search in current folder
    csv_files = [f for f in os.listdir('.') if f.startswith('s3_access_log_') and f.endswith('.csv')]
    if csv_files:
        return max(csv_files, key=os.path.getctime)
    
    return None


def run_s3_access_parser():
    """Run S3 Access Log parser"""
    print("Starting S3 Access Log parsing...")
    
    try:
        # S3 Access Log parser import and execution
        from s3_access_log_parser import S3AccessLogParser
        
        # Default path setting
        input_path = r"C:\Users\kelly.jang\Desktop\새 폴더\AWS_LOG\S3 Server Access Log (S3)"
        
        # Directory structure creation
        result_dir = "result"
        parse_log_dir = os.path.join(result_dir, "parse_log")
        
        # Output directory creation
        os.makedirs(parse_log_dir, exist_ok=True)
        
        # Parser execution
        parser = S3AccessLogParser()
        events = parser.parse_directory(input_path)
        
        if events:
            # Output filename generation
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(parse_log_dir, f"s3_access_log_{timestamp}.csv")
            
            # CSV save
            parser.save_to_csv(events, output_file)
            print(f"[SUCCESS] S3 Access Log parsing completed: {output_file}")
            return output_file
        else:
            print("[ERROR] No events parsed.")
            return None
            
    except Exception as e:
        print(f"[ERROR] Error during S3 Access Log parsing: {e}")
        return None


def main():
    """Main function"""
    print("[INFO] S3 Access Log Analyzer")
    print("=" * 50)
    
    # Directory structure creation
    result_dir = "result"
    parse_log_dir = os.path.join(result_dir, "parse_log")
    analysis_dir = os.path.join(result_dir, "analysis")
    report_dir = os.path.join(result_dir, "report")
    
    # Directory creation
    os.makedirs(parse_log_dir, exist_ok=True)
    os.makedirs(analysis_dir, exist_ok=True)
    os.makedirs(report_dir, exist_ok=True)
    
    # 1. S3 Access Log CSV file search
    latest_csv = find_s3_access_csv()
    
    if latest_csv is None:
        print("No S3 Access Log CSV file found. Starting parsing...")
        latest_csv = run_s3_access_parser()
        
        if latest_csv is None:
            print("[ERROR] Failed to parse S3 Access Log. Exiting.")
            return
    
    print(f"Analysis file: {latest_csv}")
    
    # 2. Analysis execution
    print(f"\n[INFO] Starting S3 Access Log analysis...")
    analyzer = S3AccessLogAnalyzer(latest_csv)
    
    # 3. Excel file generation
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = os.path.join(analysis_dir, f"s3_access_analysis_{timestamp}.xlsx")
    analysis_data = analyzer.save_analysis_to_excel(excel_file)
    
    # 4. HTML report generation
    html_file = os.path.join(report_dir, f"s3_access_analysis_{timestamp}.html")
    analyzer.generate_html_report(analysis_data, html_file)
    
    print(f"\n[COMPLETE] S3 Access Log analysis completed!")
    print(f"[INFO] Excel file: {excel_file}")
    print(f"[INFO] HTML report: {html_file}")
    print(f"[INFO] Check the detailed analysis results with 5 sheets and HTML report!")
    
    return excel_file, html_file


def run_analysis(csv_file: str, analysis_dir: str, report_dir: str) -> tuple:
    """Run complete S3 Server Access Log analysis and return file paths"""
    try:
        # Create directories
        os.makedirs(analysis_dir, exist_ok=True)
        os.makedirs(report_dir, exist_ok=True)
        
        # Create analyzer instance with CSV file path
        analyzer = S3AccessLogAnalyzer(csv_file)
        
        # Generate output filename for Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"s3_access_analysis_{timestamp}.xlsx"
        excel_file = os.path.join(analysis_dir, excel_filename)
        
        # Run analysis and save to Excel
        analysis_data = analyzer.save_analysis_to_excel(excel_file)
        
        # Generate HTML report
        html_filename = f"s3_access_analysis_{timestamp}.html"
        html_file = os.path.join(report_dir, html_filename)
        analyzer.generate_html_report(analysis_data, html_file)
        
        return excel_file, html_file
        
    except Exception as e:
        print(f"[ERROR] S3 Server Access Log analysis failed: {str(e)}")
        raise


if __name__ == "__main__":
    main()
