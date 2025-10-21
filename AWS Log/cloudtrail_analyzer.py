# -*- coding: utf-8 -*-
"""
CloudTrail Log Analyzer - Generate Excel file with 7 sheets
"""

import pandas as pd
import numpy as np
import sys
import json
import re
from datetime import datetime, timezone, timedelta
from collections import defaultdict
import os

# Windows Korean encoding issue resolution
if sys.platform.startswith('win'):
    os.environ['PYTHONIOENCODING'] = 'utf-8'
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    else:
        import codecs
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.detach())
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.detach())


class CloudTrailAnalyzer:
    def __init__(self, csv_file, mitre_config_path='mitre_config.json'):
        """Class for analyzing CloudTrail CSV files"""
        self.df = pd.read_csv(csv_file)
        print(f"Loaded data: {len(self.df)} events")
        
        # Load Mitre ATT&CK configuration
        self.mitre_config = self._load_mitre_config(mitre_config_path)
        
        # Convert time columns to datetime
        self._prepare_data()
    
    def _load_mitre_config(self, config_path):
        """Load Mitre ATT&CK configuration file"""
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except FileNotFoundError:
            print(f"[WARNING] Mitre config file not found: {config_path}")
            return {}
        except Exception as e:
            print(f"[WARNING] Mitre config file load error: {e}")
            return {}
    
    def _prepare_data(self):
        """Data preprocessing"""
        # Convert time columns to datetime and remove timezone info for Excel compatibility
        if 'eventTime' in self.df.columns:
            self.df['eventTime'] = pd.to_datetime(self.df['eventTime'], errors='coerce')
            # Remove timezone info for Excel compatibility
            self.df['eventTime'] = self.df['eventTime'].dt.tz_localize(None)
        
        if 'eventTimeLocal' in self.df.columns:
            # eventTimeLocal is already converted to local time by region in parser
            # Convert to datetime and ensure it's timezone-naive for Excel compatibility
            self.df['eventTimeLocal'] = pd.to_datetime(self.df['eventTimeLocal'], errors='coerce')
            # Remove timezone info for Excel compatibility
            try:
                self.df['eventTimeLocal'] = self.df['eventTimeLocal'].dt.tz_localize(None)
            except:
                # If already timezone-naive, continue
                pass
        
        # Mitre ATT&CK tactics mapping
        self.df['Mitre ATT&CK'] = self.df['eventName'].apply(self._map_mitre_tactics)
    
    def _map_mitre_tactics(self, event_name):
        """Map EventName to Mitre ATT&CK tactics (CloudTrail specific)"""
        if not event_name or not self.mitre_config:
            return ''
        
        # CloudTrail specific tactics
        cloudtrail_config = self.mitre_config.get('cloudtrail', {})
        for tactics, events in cloudtrail_config.items():
            if event_name in events:
                return tactics
        return ''
    
    def analyze_top_ips(self, top_n=20):
        """1. Top 20 IPs by first access time, last access time, and access count"""
        print(f"\n1. Analyzing top {top_n} IPs...")
        
        # Calculate IP statistics
        ip_stats = []
        
        for ip in self.df['sourceIpAddress'].value_counts().head(top_n).index:
            if pd.isna(ip) or ip == '':
                continue
                
            ip_data = self.df[self.df['sourceIpAddress'] == ip]
            
            # First/last access time
            first_access_utc = ip_data['eventTime'].min()
            last_access_utc = ip_data['eventTime'].max()
            first_access_local = ip_data['eventTimeLocal'].min()
            last_access_local = ip_data['eventTimeLocal'].max()
            
            ip_stats.append({
                'IP': ip,
                'First Access Time (UTC+0)': first_access_utc.strftime('%Y-%m-%d %H:%M:%S') if pd.notna(first_access_utc) else '',
                'First Access Time (Local)': first_access_local.strftime('%Y-%m-%d %H:%M:%S') if pd.notna(first_access_local) else '',
                'Last Access Time (UTC+0)': last_access_utc.strftime('%Y-%m-%d %H:%M:%S') if pd.notna(last_access_utc) else '',
                'Last Access Time (Local)': last_access_local.strftime('%Y-%m-%d %H:%M:%S') if pd.notna(last_access_local) else '',
                'Access Count': len(ip_data)
            })
        
        return pd.DataFrame(ip_stats)
    
    def analyze_Nightshift_events(self):
        """2. EventName statistics for nightshift hours (22:00 ~ 06:00) - all events"""
        print(f"\n2. Analyzing Nightshift EventNames (22:00-06:00)...")
        
        # Filter nightshift hours based on local time
        if 'eventTimeLocal' in self.df.columns:
            self.df['hour'] = self.df['eventTimeLocal'].dt.hour
            nightshift_mask = (self.df['hour'] >= 22) | (self.df['hour'] <= 6)
            nightshift_events = self.df[nightshift_mask]
        else:
            # Use UTC time if local time is not available
            self.df['hour'] = self.df['eventTime'].dt.hour
            nightshift_mask = (self.df['hour'] >= 22) | (self.df['hour'] <= 6)
            nightshift_events = self.df[nightshift_mask]
        
        # EventName statistics (all events)
        event_stats = nightshift_events['eventName'].value_counts()
        
        result = []
        for event_name, count in event_stats.items():
            result.append({
                'EventName': event_name,
                'Nightshift Event Count': count,
                'Event Statistics (%)': round((count / len(nightshift_events)) * 100, 2)
            })
        
        return pd.DataFrame(result)
    
    def analyze_nightshift_ips(self):
        """2. Statistics for sourceIPs that generated events during nightshift hours (22:00-06:00)"""
        print(f"\n2. Analyzing nightshift IPs (22:00-06:00)...")
        
        # Check if eventTimeLocal column exists
        if 'eventTimeLocal' not in self.df.columns:
            return pd.DataFrame({'IP': ['No local time data'], 'Access Count': [0]})
        
        # Convert eventTimeLocal to datetime
        self.df['eventTimeLocal'] = pd.to_datetime(self.df['eventTimeLocal'], errors='coerce')
        
        # Filter nightshift hours (22:00-06:00)
        nightshift_mask = (
            (self.df['eventTimeLocal'].dt.hour >= 22) | 
            (self.df['eventTimeLocal'].dt.hour < 6)
        )
        nightshift_events = self.df[nightshift_mask]
        
        if len(nightshift_events) == 0:
            return pd.DataFrame({'IP': ['No nightshift events'], 'Access Count': [0]})
        
        # Statistics by IP
        ip_stats = []
        for ip in nightshift_events['sourceIpAddress'].unique():
            if pd.notna(ip):
                ip_data = nightshift_events[nightshift_events['sourceIpAddress'] == ip]
                
                ip_stats.append({
                    'IP': ip,
                    'Access Count': len(ip_data)
                })
        
        # Sort by Access Count in descending order
        ip_stats_df = pd.DataFrame(ip_stats)
        if not ip_stats_df.empty:
            ip_stats_df = ip_stats_df.sort_values('Access Count', ascending=False)
        
        return ip_stats_df
    
    def analyze_event_names(self):
        """3. EventName statistics"""
        print(f"\n3. Analyzing EventName statistics...")
        
        event_stats = self.df['eventName'].value_counts()
        
        result = []
        for event_name, count in event_stats.items():
            result.append({
                'EventName': event_name,
                'Event Count': count,
                'Event Statistics (%)': round((count / len(self.df)) * 100, 2)
            })
        
        return pd.DataFrame(result)
    
    def analyze_user_agents(self):
        """4. UserAgent statistics - detailed classification"""
        print(f"\n4. Analyzing UserAgent statistics...")
        
        # Detailed UserAgent classification
        def classify_user_agent_detailed(ua):
            if pd.isna(ua) or ua == '':
                return 'Unknown/Empty'
            
            ua_lower = str(ua).lower()
            
            # AWS CLI/SDK detailed classification
            if 'aws-cli' in ua_lower:
                if 'python' in ua_lower:
                    return 'AWS CLI (Python)'
                elif 'java' in ua_lower:
                    return 'AWS CLI (Java)'
                else:
                    return 'AWS CLI (Other)'
            elif 'aws-sdk' in ua_lower:
                if 'python' in ua_lower:
                    return 'AWS SDK (Python)'
                elif 'java' in ua_lower:
                    return 'AWS SDK (Java)'
                elif 'javascript' in ua_lower or 'node' in ua_lower:
                    return 'AWS SDK (JavaScript/Node.js)'
                elif 'go' in ua_lower:
                    return 'AWS SDK (Go)'
                elif 'php' in ua_lower:
                    return 'AWS SDK (PHP)'
                elif 'ruby' in ua_lower:
                    return 'AWS SDK (Ruby)'
                elif 'dotnet' in ua_lower or 'c#' in ua_lower:
                    return 'AWS SDK (.NET)'
                else:
                    return 'AWS SDK (Other)'
            
            # Browser detailed classification
            elif 'mozilla' in ua_lower:
                if 'firefox' in ua_lower:
                    return 'Browser (Firefox)'
                elif 'chrome' in ua_lower:
                    return 'Browser (Chrome)'
                else:
                    return 'Browser (Mozilla-based)'
            elif 'chrome' in ua_lower:
                return 'Browser (Chrome)'
            elif 'safari' in ua_lower and 'chrome' not in ua_lower:
                return 'Browser (Safari)'
            elif 'edge' in ua_lower:
                return 'Browser (Edge)'
            elif 'opera' in ua_lower:
                return 'Browser (Opera)'
            
            # AWS Console
            elif 'console' in ua_lower or 'aws-console' in ua_lower:
                return 'AWS Console'
            
            # Programming languages/frameworks
            elif 'python' in ua_lower:
                if 'requests' in ua_lower:
                    return 'Python (requests)'
                elif 'boto3' in ua_lower:
                    return 'Python (boto3)'
                else:
                    return 'Python (Other)'
            elif 'java' in ua_lower:
                return 'Java Application'
            elif 'go' in ua_lower:
                return 'Go Application'
            elif 'node' in ua_lower or 'javascript' in ua_lower:
                return 'Node.js/JavaScript'
            elif 'php' in ua_lower:
                return 'PHP Application'
            elif 'ruby' in ua_lower:
                return 'Ruby Application'
            elif 'curl' in ua_lower:
                return 'cURL'
            elif 'wget' in ua_lower:
                return 'wget'
            elif 'postman' in ua_lower:
                return 'Postman'
            elif 'insomnia' in ua_lower:
                return 'Insomnia'
            
            # Mobile/tablet
            elif 'mobile' in ua_lower or 'android' in ua_lower:
                return 'Mobile (Android)'
            elif 'iphone' in ua_lower or 'ipad' in ua_lower:
                return 'Mobile (iOS)'
            
            # Others
            elif 'bot' in ua_lower or 'crawler' in ua_lower or 'spider' in ua_lower:
                return 'Bot/Crawler'
            elif 'aws' in ua_lower:
                return 'AWS Service'
            else:
                return 'Other/Unknown'
        
        self.df['UserAgent_Detailed'] = self.df['userAgent'].apply(classify_user_agent_detailed)
        
        # Calculate statistics
        ua_stats = self.df['UserAgent_Detailed'].value_counts()
        
        result = []
        for ua_type, count in ua_stats.items():
            result.append({
                'UserAgent Type': ua_type,
                'Usage Count': count,
                'Usage Statistics (%)': round((count / len(self.df)) * 100, 2)
            })
        
        return pd.DataFrame(result)
    
    def analyze_failed_auth(self):
        """5. Failed authentication/authorization statistics - brute force attack detection"""
        print(f"\n5. Analyzing failed authentication/authorization...")
        
        # Failure type classification (brute force attack detection perspective)
        failure_categories = {
            'Authentication Failures': [
                'InvalidUserID', 'InvalidUserID.NotFound', 'AuthenticationFailed', 
                'LoginFailed', 'InvalidToken', 'ExpiredToken', 'TokenRefreshRequired'
            ],
            'Authorization Failures': [
                'AccessDenied', 'AccessDeniedException', 'Forbidden', 'ForbiddenException',
                'UnauthorizedOperation', 'UnauthorizedOperationException'
            ],
            'API Key/Signature Errors': [
                'InvalidAccessKeyId', 'InvalidAccessKeyId.NotFound', 'SignatureDoesNotMatch'
            ]
        }
        
        # Filter events by each failure type
        failed_events_by_category = {}
        all_failed_mask = pd.Series([False] * len(self.df), index=self.df.index)
        
        for category, keywords in failure_categories.items():
            category_mask = pd.Series([False] * len(self.df), index=self.df.index)
            for col in ['errorCode', 'errorMessage', 'eventName']:
                if col in self.df.columns:
                    for keyword in keywords:
                        category_mask |= self.df[col].astype(str).str.contains(keyword, case=False, na=False)
            
            failed_events_by_category[category] = self.df[category_mask]
            all_failed_mask = all_failed_mask | category_mask
        
        all_failed_events = self.df[all_failed_mask]
        
        if len(all_failed_events) == 0:
            return pd.DataFrame({'Failure Type': ['No failures found'], 'Count': [0]})
        
        # Detailed statistics by failure type (separated by IP and Users)
        result = []
        for category, failed_events in failed_events_by_category.items():
            if len(failed_events) > 0:
                # IP statistics
                unique_ips = failed_events['sourceIpAddress'].nunique() if 'sourceIpAddress' in failed_events.columns else 0
                top_ips = []
                if 'sourceIpAddress' in failed_events.columns:
                    ip_counts = failed_events['sourceIpAddress'].value_counts()
                    top_ips = [f"{ip} ({count})" for ip, count in ip_counts.items()]
                
                # UserName statistics (including Root User)
                unique_users = 0
                user_series = None
                
                if 'userIdentity.userName' in failed_events.columns:
                    user_series = failed_events['userIdentity.userName']
                elif 'userIdentity.principalId' in failed_events.columns:
                    user_series = failed_events['userIdentity.principalId']
                elif 'userIdentity.type' in failed_events.columns:
                    # Root User may have type 'Root'
                    user_series = failed_events['userIdentity.type']
                
                if user_series is not None:
                    unique_users = user_series.nunique()
                
                top_users = []
                if user_series is not None:
                    user_counts = user_series.value_counts()
                    top_users = [f"{user} ({count})" for user, count in user_counts.items() if pd.notna(user)]
                
                # Add IP information row
                result.append({
                    'Category': f"{category} - IPs",
                    'Count': unique_ips,
                    'Details': '; '.join(top_ips) if top_ips else 'N/A'
                })
                
                # Add Users information row
                result.append({
                    'Category': f"{category} - Users",
                    'Count': unique_users,
                    'Details': '; '.join(top_users) if top_users else 'N/A'
                })
        
        return pd.DataFrame(result)
    
    def get_failed_auth_events_by_category(self):
        """Return authentication/authorization failure events separated by failure type"""
        print(f"\n6. Extracting failed authentication events by category...")
        
        # Failure type classification
        failure_categories = {
            'Authentication_Failures': [
                'InvalidUserID', 'InvalidUserID.NotFound', 'AuthenticationFailed', 
                'LoginFailed', 'InvalidToken', 'ExpiredToken', 'TokenRefreshRequired'
            ],
            'Authorization_Failures': [
                'AccessDenied', 'AccessDeniedException', 'Forbidden', 'ForbiddenException',
                'UnauthorizedOperation', 'UnauthorizedOperationException'
            ],
            'API_Key_Errors': [
                'InvalidAccessKeyId', 'InvalidAccessKeyId.NotFound', 'SignatureDoesNotMatch'
            ]
        }
        
        failed_events_by_category = {}
        
        for category, keywords in failure_categories.items():
            category_mask = False
            for col in ['errorCode', 'errorMessage', 'eventName']:
                if col in self.df.columns:
                    for keyword in keywords:
                        category_mask |= self.df[col].astype(str).str.contains(keyword, case=False, na=False)
            
            if category_mask.any():
                failed_events = self.df[category_mask].copy()
                
                # Additional columns for brute force attack detection
                failed_events['Failure_Time'] = failed_events['eventTime']
                failed_events['Failure_IP'] = failed_events['sourceIpAddress']
                failed_events['Failure_User'] = failed_events['userIdentity.userName']
                failed_events['Failure_Account'] = failed_events['userIdentity.accountId']
                failed_events['Failure_Type'] = failed_events['errorCode']
                failed_events['Failure_Message'] = failed_events['errorMessage']
                
                # Adjust column order (important columns first)
                priority_cols = ['Failure_Time', 'Failure_IP', 'Failure_User', 'Failure_Account', 'Failure_Type', 'Failure_Message', 'eventName']
                other_cols = [col for col in failed_events.columns if col not in priority_cols]
                failed_events = failed_events[priority_cols + other_cols]
                
                # Convert timezone-aware datetime columns to strings
                for col in failed_events.columns:
                    if failed_events[col].dtype.name.startswith('datetime'):
                        failed_events[col] = failed_events[col].astype(str)
                
                failed_events_by_category[category] = failed_events
        
        return failed_events_by_category
    
    def analyze_regions(self):
        """6. Region-based statistics"""
        print(f"\n6. Analyzing region statistics...")
        
        region_stats = self.df['awsRegion'].value_counts()
        
        result = []
        for region, count in region_stats.items():
            result.append({
                'Region': region,
                'Event Count': count,
                'Event Statistics (%)': round((count / len(self.df)) * 100, 2)
            })
        
        return pd.DataFrame(result)
    
    def analyze_console_login(self, days_filter=None):
        """8. ConsoleLogin history statistics analysis"""
        print(f"\n8. Analyzing ConsoleLogin statistics...")
        
        # Filter only ConsoleLogin events
        console_login_events = self.df[self.df['eventName'] == 'ConsoleLogin']
        
        if len(console_login_events) == 0:
            return pd.DataFrame({
                'User': ['No ConsoleLogin events found'],
                'Event Time (UTC)': [''],
                'Event Time (Local)': [''],
                'Source IP': [''],
                'Region': [''],
                'Login Status': [''],
                'User Agent': ['']
            })
        
        # Apply date filter (last N days)
        if days_filter:
            from datetime import datetime, timedelta
            cutoff_date = datetime.now() - timedelta(days=days_filter)
            console_login_events = console_login_events[
                pd.to_datetime(console_login_events['eventTime'], errors='coerce') >= cutoff_date
            ]
            print(f"Filtering ConsoleLogin events to last {days_filter} days")
        
        # Detailed information for each login event
        login_details = []
        for _, event in console_login_events.iterrows():
            # Skip if user is empty or 'Unknown'
            user_name = event.get('userIdentity.userName', '')
            if pd.isna(user_name) or user_name == '' or user_name == 'Unknown':
                continue
                
            # Determine login success/failure (failed if errorCode exists)
            login_status = 'Success' if pd.isna(event.get('errorCode')) or event.get('errorCode') == '' else 'Failed'
            
            login_details.append({
                'User': user_name,
                'Event Time (UTC)': event.get('eventTime', ''),
                'Event Time (Local)': event.get('eventTimeLocal', ''),
                'Source IP': event.get('sourceIpAddress', ''),
                'Region': event.get('awsRegion', ''),
                'Login Status': login_status,
                'User Agent': event.get('userAgent', '')
            })
        
        # Sort by time
        login_details_df = pd.DataFrame(login_details)
        login_details_df = login_details_df.sort_values('Event Time (UTC)', ascending=False)
        
        return login_details_df
    
    def analyze_user_management(self):
        """9. CreateUser detailed history analysis"""
        print(f"\n9. Analyzing CreateUser events...")
        
        # Filter only CreateUser events
        create_user_events = self.df[self.df['eventName'] == 'CreateUser']
        
        if len(create_user_events) == 0:
            return pd.DataFrame({
                'Created User': ['No CreateUser events found'],
                'Created By (ARN)': [''],
                'Event Time (UTC)': [''],
                'Event Time (Local)': [''],
                'Source IP': [''],
                'Region': ['']
            })
        
        # Detailed information for each CreateUser event
        user_creation_details = []
        for _, event in create_user_events.iterrows():
            # Extract created user name (userName from requestElements)
            created_user = 'Unknown'
            if 'requestElements' in event and event['requestElements']:
                try:
                    import json
                    request_elements = json.loads(event['requestElements']) if isinstance(event['requestElements'], str) else event['requestElements']
                    created_user = request_elements.get('userName', 'Unknown')
                except:
                    created_user = 'Unknown'
            
            # ARN of the creator
            creator_arn = event.get('userIdentity.arn', 'Unknown')
            
            user_creation_details.append({
                'Created User': created_user,
                'Created By (ARN)': creator_arn,
                'Event Time (UTC)': event.get('eventTime', ''),
                'Event Time (Local)': event.get('eventTimeLocal', ''),
                'Source IP': event.get('sourceIpAddress', ''),
                'Region': event.get('awsRegion', '')
            })
        
        # Sort by time
        user_creation_df = pd.DataFrame(user_creation_details)
        user_creation_df = user_creation_df.sort_values('Event Time (UTC)', ascending=False)
        
        return user_creation_df
    
    def analyze_mitre_attack(self):
        """7. Mitre ATT&CK tactics-based analysis - detailed EventName statistics by tactics"""
        print(f"\n7. Analyzing Mitre ATT&CK tactics...")
        
        # Filter only mapped tactics (exclude empty strings)
        mapped_events = self.df[self.df['Mitre ATT&CK'] != '']
        
        if len(mapped_events) == 0:
            return pd.DataFrame({'Mitre ATT&CK Tactics': ['No mapped tactics found'], 'EventName': [''], 'EventName Hit Count': [0], 'Event Statistics (%)': [0]})
        
        # Generate EventName statistics by tactics
        result_data = []
        
        for tactics in mapped_events['Mitre ATT&CK'].unique():
            if tactics:  # Exclude empty strings
                tactics_events = mapped_events[mapped_events['Mitre ATT&CK'] == tactics]
                event_name_stats = tactics_events['eventName'].value_counts()
                
                # Generate rows for each EventName
                for event_name, count in event_name_stats.items():
                    result_data.append({
                        'Mitre ATT&CK Tactics': tactics,
                        'EventName': event_name,
                        'EventName Hit Count': count,
                        'Event Statistics (%)': round((count / len(self.df)) * 100, 2)
                    })
        
        # Sort by EventName Hit Count in descending order
        result_df = pd.DataFrame(result_data)
        if not result_df.empty:
            result_df = result_df.sort_values(['Mitre ATT&CK Tactics', 'EventName Hit Count'], ascending=[True, False])
        
        return result_df
    
    def get_mitre_attack_events(self):
        """Integrate all Mitre ATT&CK tactics events into one sheet"""
        print(f"\n8. Extracting Mitre ATT&CK tactics events...")
        
        # Filter only mapped tactics (exclude empty strings)
        mapped_events = self.df[self.df['Mitre ATT&CK'] != ''].copy()
        
        if len(mapped_events) == 0:
            return pd.DataFrame({'Mitre ATT&CK': ['No mapped tactics found'], 'eventName': [''], 'eventTime': [''], 'sourceIpAddress': ['']})
        
        # Move Mitre ATT&CK column to the front
        cols = list(mapped_events.columns)
        if 'Mitre ATT&CK' in cols:
            cols.remove('Mitre ATT&CK')
            cols.insert(0, 'Mitre ATT&CK')
            mapped_events = mapped_events[cols]
        
        # Convert timezone-aware datetime columns to strings
        for col in mapped_events.columns:
            if mapped_events[col].dtype.name.startswith('datetime'):
                mapped_events[col] = mapped_events[col].astype(str)
        
        # Sort by Mitre ATT&CK tactics
        mapped_events = mapped_events.sort_values(['Mitre ATT&CK', 'eventTime'], ascending=[True, True])
        
        return mapped_events
    
    def generate_html_report(self, report_dir, analysis_data=None):
        """Generate HTML report"""
        print(f"\n[INFO] Generating HTML report...")
        
        if analysis_data is None:
            # 1. Top 20 IP Statistics
            top_ips = self.analyze_top_ips(20)
            
            # 2. Nightshift IP Statistics
            nightshift_ips = self.analyze_nightshift_ips()
            
            # 3. Nightshift EventName Statistics (Top 20)
            nightshift_events = self.analyze_Nightshift_events()
            nightshift_events_top20 = nightshift_events.head(20)
            
            # 4. EventName Statistics (Top 20)
            event_names = self.analyze_event_names()
            event_names_top20 = event_names.head(20)
            
            # 5. UserAgent Statistics
            user_agents = self.analyze_user_agents()
            
            # 6. Failed Auth Statistics
            failed_auth = self.analyze_failed_auth()
            
            # 7. Region Statistics
            regions = self.analyze_regions()
            
            # 8. ConsoleLogin history statistics (for HTML - last 30 days)
            console_login = self.analyze_console_login(days_filter=30)
            
            # 9. User Management statistics
            user_management = self.analyze_user_management()
            
            # 10. Mitre ATT&CK Statistics
            mitre_attack = self.analyze_mitre_attack()
        else:
            # Use already computed data
            top_ips = analysis_data['top_ips']
            nightshift_ips = analysis_data['nightshift_ips']
            nightshift_events_top20 = analysis_data['nightshift_events'].head(20)
            event_names_top20 = analysis_data['event_names'].head(20)
            user_agents = analysis_data['user_agents']
            failed_auth = analysis_data['failed_auth']
            regions = analysis_data['regions']
            console_login = analysis_data.get('console_login', self.analyze_console_login())
            # For HTML, apply 30-day filter
            if 'console_login' in analysis_data:
                console_login = self.analyze_console_login(days_filter=30)
            user_management = analysis_data.get('user_management', self.analyze_user_management())
            mitre_attack = analysis_data['mitre_attack']
        
        # Generate HTML
        html_content = self._create_html_content(
            top_ips, nightshift_ips, nightshift_events_top20, event_names_top20, 
            user_agents, failed_auth, regions, console_login, user_management, mitre_attack
        )
        
        # Save HTML file (in report folder)
        os.makedirs(report_dir, exist_ok=True)
        
        html_filename = "cloudtrail_analysis_report.html"
        html_file = os.path.join(report_dir, html_filename)
        
        with open(html_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"[SUCCESS] HTML report created: {html_file}")
        return html_file
    
    def _create_html_content(self, top_ips, nightshift_ips, nightshift_events, event_names, 
                           user_agents, failed_auth, regions, console_login, user_management, mitre_attack):
        """Create HTML content"""
        
        html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Summary Analysis Report (CloudTrail Log)</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
            color: #333;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 10px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}
        .header h1 {{
            margin: 0;
            font-size: 2.5em;
            font-weight: 300;
        }}
        .header p {{
            margin: 10px 0 0 0;
            opacity: 0.9;
            font-size: 1.1em;
        }}
        .section {{
            padding: 30px;
            border-bottom: 1px solid #eee;
        }}
        .section:last-child {{
            border-bottom: none;
        }}
        .section h2 {{
            color: #2c3e50;
            margin: 0 0 20px 0;
            font-size: 1.8em;
            border-left: 4px solid #3498db;
            padding-left: 15px;
        }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 20px;
            margin-bottom: 20px;
        }}
        .stat-card {{
            background: #f8f9fa;
            border-radius: 8px;
            padding: 20px;
            border-left: 4px solid #3498db;
        }}
        .stat-card h3 {{
            margin: 0 0 10px 0;
            color: #2c3e50;
            font-size: 1.2em;
        }}
        .stat-value {{
            font-size: 2em;
            font-weight: bold;
            color: #3498db;
            margin: 10px 0;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        th {{
            background: #34495e;
            color: white;
            padding: 15px;
            text-align: left;
            font-weight: 600;
        }}
        td {{
            padding: 12px 15px;
            border-bottom: 1px solid #eee;
        }}
        tr:nth-child(even) {{
            background: #f8f9fa;
        }}
        tr:hover {{
            background: #e3f2fd;
        }}
        .chart-container {{
            margin: 20px 0;
            text-align: center;
        }}
        .no-data {{
            text-align: center;
            color: #7f8c8d;
            font-style: italic;
            padding: 40px;
        }}
        .footer {{
            background: #2c3e50;
            color: white;
            text-align: center;
            padding: 20px;
            font-size: 0.9em;
        }}
        .chart-container {{
            margin: 20px 0;
            height: 400px;
            position: relative;
        }}
        .chart-wrapper {{
            width: 100%;
            height: 400px;
            margin: 20px 0;
            background: #fafafa;
            border-radius: 8px;
            padding: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
    </style>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Summary Analysis Report (CloudTrail Log)</h1>
            <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
        
        {self._create_top_ips_section(top_ips)}
        {self._create_nightshift_ips_section(nightshift_ips)}
        {self._create_nightshift_events_section(nightshift_events)}
        {self._create_event_names_section(event_names)}
        {self._create_user_agents_section(user_agents)}
        {self._create_failed_auth_section(failed_auth)}
        {self._create_regions_section(regions)}
        {self._create_console_login_section(console_login)}
        {self._create_user_management_section(user_management)}
        {self._create_mitre_attack_section(mitre_attack)}
        
        <div class="footer">
            <p>Summary Analysis Report (CloudTrail Log) - Generated by PLAINBIT Co., LTD.</p>
        </div>
    </div>
</body>
</html>
        """
        return html
    
    def _create_top_ips_section(self, top_ips):
        """Create Top IPs section"""
        if top_ips.empty:
            return f"""
        <div class="section">
            <h2>1. Access IP Statistics (Top 20)</h2>
            <div class="no-data">No IP data available</div>
        </div>
            """
        
        # Prepare chart data
        chart_data = top_ips.head(10)  # Show only top 10 in chart
        chart_labels = [str(ip) for ip in chart_data['IP']]
        chart_values = chart_data['Access Count'].tolist()
        
        table_rows = ""
        for _, row in top_ips.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{row['IP']}</td>
                <td>{row['First Access Time (UTC+0)']}</td>
                <td>{row['First Access Time (Local)']}</td>
                <td>{row['Last Access Time (UTC+0)']}</td>
                <td>{row['Last Access Time (Local)']}</td>
                <td>{row['Access Count']}</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>1. Access IP Statistics (Top 20)</h2>
            <div class="chart-wrapper">
                <canvas id="topIpsChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>IP Address</th>
                        <th>First Access (UTC+0)</th>
                        <th>First Access (Local)</th>
                        <th>Last Access (UTC+0)</th>
                        <th>Last Access (Local)</th>
                        <th>Access Count</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const topIpsCtx = document.getElementById('topIpsChart').getContext('2d');
                new Chart(topIpsCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)',
                                'rgba(39, 174, 96, 0.8)',
                                'rgba(41, 128, 185, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Top 10 IPs by Access Count'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def _create_nightshift_ips_section(self, nightshift_ips):
        """Create Nightshift IPs section"""
        if nightshift_ips.empty:
            return f"""
        <div class="section">
            <h2>2. Nightshift IP Statistics (22:00-06:00)</h2>
            <div class="no-data">No nightshift IP data available</div>
        </div>
            """
        
        # Prepare chart data
        chart_data = nightshift_ips.head(10)  # Show only top 10 in chart
        chart_labels = [str(ip) for ip in chart_data['IP']]
        chart_values = chart_data['Access Count'].tolist()
        
        table_rows = ""
        for _, row in nightshift_ips.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{row['IP']}</td>
                <td>{row['Access Count']}</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>2. Nightshift(22:00-06:00) IP Statistics</h2>
            <div class="chart-wrapper">
                <canvas id="nightshiftIpsChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>IP Address</th>
                        <th>Access Count</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const nightshiftIpsCtx = document.getElementById('nightshiftIpsChart').getContext('2d');
                new Chart(nightshiftIpsCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)',
                                'rgba(39, 174, 96, 0.8)',
                                'rgba(41, 128, 185, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Top 10 Nightshift IPs by Access Count'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def _create_nightshift_events_section(self, nightshift_events):
        """Create Nightshift Events section"""
        if nightshift_events.empty:
            return f"""
        <div class="section">
            <h2>3. Nightshift EventName Statistics (Top 20)</h2>
            <div class="no-data">No nightshift events found</div>
        </div>
            """
        
        # Prepare chart data
        chart_data = nightshift_events.head(10)  # Show only top 10 in chart
        chart_labels = [str(event) for event in chart_data['EventName']]
        chart_values = chart_data['Nightshift Event Count'].tolist()
        
        table_rows = ""
        for _, row in nightshift_events.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{row['EventName']}</td>
                <td>{row['Nightshift Event Count']}</td>
                <td>{row['Event Statistics (%)']}%</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>2. Nightshift EventName Statistics (Top 20)</h2>
            <div class="chart-wrapper">
                <canvas id="nightshiftEventsChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>Event Name</th>
                        <th>Nightshift Event Count</th>
                        <th>Event Statistics (%)</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const nightshiftEventsCtx = document.getElementById('nightshiftEventsChart').getContext('2d');
                new Chart(nightshiftEventsCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)',
                                'rgba(39, 174, 96, 0.8)',
                                'rgba(41, 128, 185, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Top 10 Nightshift Events Distribution'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def _create_event_names_section(self, event_names):
        """Create Event Names section"""
        if event_names.empty:
            return f"""
        <div class="section">
            <h2>4. EventName Statistics (Top 20)</h2>
            <div class="no-data">No event data available</div>
        </div>
            """
        
        # Prepare chart data
        chart_data = event_names.head(10)  # Show only top 10 in chart
        chart_labels = [str(event) for event in chart_data['EventName']]
        chart_values = chart_data['Event Count'].tolist()
        
        table_rows = ""
        for _, row in event_names.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{row['EventName']}</td>
                <td>{row['Event Count']}</td>
                <td>{row['Event Statistics (%)']}%</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>4. EventName Statistics (Top 20)</h2>
            <div class="chart-wrapper">
                <canvas id="eventNamesChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>Event Name</th>
                        <th>Event Count</th>
                        <th>Event Statistics (%)</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const eventNamesCtx = document.getElementById('eventNamesChart').getContext('2d');
                new Chart(eventNamesCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)',
                                'rgba(39, 174, 96, 0.8)',
                                'rgba(41, 128, 185, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Top 10 Event Names by Count'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def _create_user_agents_section(self, user_agents):
        """Create User Agents section"""
        if user_agents.empty:
            return f"""
        <div class="section">
            <h2>5. UserAgent Statistics</h2>
            <div class="no-data">No user agent data available</div>
        </div>
            """
        
        # Prepare chart data
        chart_data = user_agents.head(8)  # Show only top 8 in chart
        chart_labels = [str(ua) for ua in chart_data['UserAgent Type']]
        chart_values = chart_data['Usage Count'].tolist()
        
        table_rows = ""
        for _, row in user_agents.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{row['UserAgent Type']}</td>
                <td>{row['Usage Count']}</td>
                <td>{row['Usage Statistics (%)']}%</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>5. UserAgent Statistics</h2>
            <div class="chart-wrapper">
                <canvas id="userAgentsChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>UserAgent Type</th>
                        <th>Usage Count</th>
                        <th>Usage Statistics (%)</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const userAgentsCtx = document.getElementById('userAgentsChart').getContext('2d');
                new Chart(userAgentsCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'UserAgent Types Distribution'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def _create_failed_auth_section(self, failed_auth):
        """Create Failed Auth section"""
        if failed_auth.empty:
            return f"""
        <div class="section">
            <h2>6. Failed Authentication/Authorization Statistics</h2>
            <div class="no-data">No failed authentication events found</div>
        </div>
            """
        
        table_rows = ""
        for _, row in failed_auth.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{row['Category']}</td>
                <td>{row['Count']}</td>
                <td>{row['Details']}</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>6. Failed Authentication/Authorization Statistics</h2>
            <table>
                <thead>
                    <tr>
                        <th>Category</th>
                        <th>Count</th>
                        <th>Details</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
        </div>
        """
    
    def _create_regions_section(self, regions):
        """Create Regions section"""
        if regions.empty:
            return f"""
        <div class="section">
            <h2>7. Region Statistics</h2>
            <div class="no-data">No region data available</div>
        </div>
            """
        
        # Prepare chart data
        chart_data = regions.head(10)  # Show only top 10 in chart
        chart_labels = [str(region) for region in chart_data['Region']]
        chart_values = chart_data['Event Count'].tolist()
        
        table_rows = ""
        for _, row in regions.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{row['Region']}</td>
                <td>{row['Event Count']}</td>
                <td>{row['Event Statistics (%)']}%</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>7. Region Statistics</h2>
            <div class="chart-wrapper">
                <canvas id="regionsChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>Region</th>
                        <th>Event Count</th>
                        <th>Event Statistics (%)</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const regionsCtx = document.getElementById('regionsChart').getContext('2d');
                new Chart(regionsCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)',
                                'rgba(39, 174, 96, 0.8)',
                                'rgba(41, 128, 185, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Top 10 Regions by Event Count'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def _create_console_login_section(self, console_login):
        """Create ConsoleLogin Statistics section"""
        if console_login.empty or 'No ConsoleLogin events found' in console_login['User'].iloc[0]:
            return f"""
        <div class="section">
            <h2>8. ConsoleLogin Statistics (Last 30 Days)</h2>
            <div class="no-data">No ConsoleLogin events found in the last 30 days</div>
        </div>
            """
        
        # Prepare chart data (login status distribution)
        status_counts = console_login['Login Status'].value_counts()
        chart_labels = status_counts.index.tolist()
        chart_values = status_counts.values.tolist()
        
        table_rows = ""
        for _, row in console_login.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{row['User']}</td>
                <td>{row['Event Time (UTC)']}</td>
                <td>{row['Event Time (Local)']}</td>
                <td>{row['Source IP']}</td>
                <td>{row['Region']}</td>
                <td>{row['Login Status']}</td>
                <td>{row['User Agent']}</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>8. ConsoleLogin Statistics (Last 30 Days)</h2>
            <div class="chart-wrapper">
                <canvas id="consoleLoginChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>User</th>
                        <th>Event Time (UTC)</th>
                        <th>Event Time (Local)</th>
                        <th>Source IP</th>
                        <th>Region</th>
                        <th>Login Status</th>
                        <th>User Agent</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const consoleLoginCtx = document.getElementById('consoleLoginChart').getContext('2d');
                new Chart(consoleLoginCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(231, 76, 60, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Login Status Distribution'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def _create_user_management_section(self, user_management):
        """Create User Management Statistics section"""
        if user_management.empty or 'No CreateUser events found' in user_management['Created User'].iloc[0]:
            return f"""
        <div class="section">
            <h2>9. CreateUser Statistics</h2>
            <div class="no-data">No CreateUser events found</div>
        </div>
            """
        
        # Prepare chart data (creator distribution)
        creator_counts = user_management['Created By (ARN)'].value_counts().head(10)
        chart_labels = [str(creator) for creator in creator_counts.index]
        chart_values = creator_counts.values.tolist()
        
        table_rows = ""
        for _, row in user_management.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{row['Created User']}</td>
                <td>{row['Created By (ARN)']}</td>
                <td>{row['Event Time (UTC)']}</td>
                <td>{row['Event Time (Local)']}</td>
                <td>{row['Source IP']}</td>
                <td>{row['Region']}</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>9. CreateUser Statistics</h2>
            <div class="chart-wrapper">
                <canvas id="userManagementChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>Created User</th>
                        <th>Created By (ARN)</th>
                        <th>Event Time (UTC)</th>
                        <th>Event Time (Local)</th>
                        <th>Source IP</th>
                        <th>Region</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const userManagementCtx = document.getElementById('userManagementChart').getContext('2d');
                new Chart(userManagementCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)',
                                'rgba(39, 174, 96, 0.8)',
                                'rgba(41, 128, 185, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Top 10 User Creators by ARN'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def _create_mitre_attack_section(self, mitre_attack):
        """Create Mitre ATT&CK section"""
        if mitre_attack.empty:
            return f"""
        <div class="section">
            <h2>10. Mitre ATT&CK Tactics Statistics</h2>
            <div class="no-data">No Mitre ATT&CK tactics found</div>
        </div>
            """
        
        # Prepare chart data (aggregated by tactics)
        tactics_counts = mitre_attack.groupby('Mitre ATT&CK Tactics')['EventName Hit Count'].sum().sort_values(ascending=False)
        chart_data = tactics_counts.head(8)  # Show only top 8 in chart
        chart_labels = [str(tactics) for tactics in chart_data.index]
        chart_values = chart_data.values.tolist()
        
        table_rows = ""
        for _, row in mitre_attack.head(20).iterrows():
            table_rows += f"""
            <tr>
                <td>{row['Mitre ATT&CK Tactics']}</td>
                <td>{row['EventName']}</td>
                <td>{row['EventName Hit Count']}</td>
                <td>{row['Event Statistics (%)']}%</td>
            </tr>
            """
        
        return f"""
        <div class="section">
            <h2>8. Mitre ATT&CK Tactics Statistics</h2>
            <div class="chart-wrapper">
                <canvas id="mitreAttackChart"></canvas>
            </div>
            <table>
                <thead>
                    <tr>
                        <th>Mitre ATT&CK Tactics</th>
                        <th>Event Name</th>
                        <th>Event Name Hit Count</th>
                        <th>Event Statistics (%)</th>
                    </tr>
                </thead>
                <tbody>
                    {table_rows}
                </tbody>
            </table>
            <script>
                const mitreAttackCtx = document.getElementById('mitreAttackChart').getContext('2d');
                new Chart(mitreAttackCtx, {{
                    type: 'doughnut',
                    data: {{
                        labels: {chart_labels},
                        datasets: [{{
                            data: {chart_values},
                            backgroundColor: [
                                'rgba(231, 76, 60, 0.8)',
                                'rgba(52, 152, 219, 0.8)',
                                'rgba(46, 204, 113, 0.8)',
                                'rgba(155, 89, 182, 0.8)',
                                'rgba(241, 196, 15, 0.8)',
                                'rgba(230, 126, 34, 0.8)',
                                'rgba(26, 188, 156, 0.8)',
                                'rgba(142, 68, 173, 0.8)'
                            ]
                        }}]
                    }},
                    options: {{
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: {{
                            title: {{
                                display: true,
                                text: 'Top 8 Mitre ATT&CK Tactics Distribution'
                            }},
                            legend: {{
                                position: 'right'
                            }}
                        }}
                    }}
                }});
            </script>
        </div>
        """
    
    def save_analysis_to_excel(self, output_file):
        """Save analysis results to Excel file"""
        print(f"\n[INFO] Creating Excel file: {output_file}")
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 1. Top 20 IP statistics
            top_ips = self.analyze_top_ips(20)
            top_ips.to_excel(writer, sheet_name='Access_IP_Statistics_(Top_20)', index=False)
            
            # 2. Nightshift IP statistics
            nightshift_ips = self.analyze_nightshift_ips()
            nightshift_ips.to_excel(writer, sheet_name='Nightshift_IP_Statistics', index=False)
            
            # 3. Nightshift EventName statistics
            nightshift_events = self.analyze_Nightshift_events()
            nightshift_events.to_excel(writer, sheet_name='Nightshift_EventName_Statistics', index=False)
            
            # 4. EventName statistics
            event_names = self.analyze_event_names()
            event_names.to_excel(writer, sheet_name='EventName_Statistics', index=False)
            
            # 5. UserAgent statistics
            user_agents = self.analyze_user_agents()
            user_agents.to_excel(writer, sheet_name='UserAgent_Statistics', index=False)
            
            # 6. Failed authentication/authorization statistics
            failed_auth = self.analyze_failed_auth()
            failed_auth.to_excel(writer, sheet_name='Failed_Auth_Statistics', index=False)
            
            # 7. Region statistics
            regions = self.analyze_regions()
            regions.to_excel(writer, sheet_name='Region_Statistics', index=False)
            
        # 8. ConsoleLogin history statistics
            console_login = self.analyze_console_login()
            console_login.to_excel(writer, sheet_name='ConsoleLogin_Statistics', index=False)
            
        # 9. CreateUser statistics
            user_management = self.analyze_user_management()
            user_management.to_excel(writer, sheet_name='CreateUser_Statistics', index=False)
            
            # 10. Mitre ATT&CK tactics statistics
            mitre_attack = self.analyze_mitre_attack()
            mitre_attack.to_excel(writer, sheet_name='Mitre_ATTACK_Tactics', index=False)
            
            # 11. Mitre ATT&CK tactics events (integrated sheet)
            mitre_events = self.get_mitre_attack_events()
            mitre_events.to_excel(writer, sheet_name='Mitre_ATTACK_Events', index=False)
        
        # Excel file formatting (header styling, column width adjustment)
        self._format_excel_sheets(output_file)
        
        print(f"[SUCCESS] Excel file created successfully: {output_file}")
        print(f"[INFO] Total sheets: 11 (11 analysis sheets)")
        
        # Return analysis data
        return {
            'top_ips': top_ips,
            'nightshift_ips': nightshift_ips,
            'nightshift_events': nightshift_events,
            'event_names': event_names,
            'user_agents': user_agents,
            'failed_auth': failed_auth,
            'regions': regions,
            'console_login': console_login,
            'user_management': user_management,
            'mitre_attack': mitre_attack,
            'mitre_events': mitre_events
        }
    
    def _format_excel_sheets(self, output_file):
        """Excel file header styling and column width adjustment"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Border, PatternFill, Font
            from openpyxl.utils import get_column_letter
            
            wb = load_workbook(output_file)
            
            # Define header row style
            header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            header_font = Font(bold=True)
            
            # Apply styling to all sheets
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Style first row (header row)
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.border = Border()  # Remove border
                    cell.fill = header_fill  # Apply background color
                    cell.font = header_font  # Apply bold font
                
                # Auto-adjust column width
                self._adjust_column_widths(ws)
            
            wb.save(output_file)
            print("[SUCCESS] Excel formatting applied successfully")
            
        except Exception as e:
            print(f"[WARNING] Error formatting Excel file: {e}")
    
    def _adjust_column_widths(self, ws):
        """Auto-adjust worksheet column width"""
        try:
            from openpyxl.utils import get_column_letter
            
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Limit minimum width 10, maximum width 50
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            print(f"[WARNING] Error adjusting column widths: {e}")


def find_cloudtrail_csv():
    """Function to find CloudTrail CSV file"""
    # Find latest CloudTrail CSV file in result/parse_log folder
    parse_log_dir = os.path.join('result', 'parse_log')
    if os.path.exists(parse_log_dir):
        csv_files = [f for f in os.listdir(parse_log_dir) if f.startswith('cloudtrail_log_') and f.endswith('.csv')]
        if csv_files:
            latest_csv = os.path.join(parse_log_dir, max(csv_files, key=lambda x: os.path.getctime(os.path.join(parse_log_dir, x))))
            return latest_csv
    
    # Find latest CloudTrail CSV file in output folder (legacy compatibility)
    output_dir = 'output'
    if os.path.exists(output_dir):
        csv_files = [f for f in os.listdir(output_dir) if f.startswith('cloudtrail_log_') and f.endswith('.csv')]
        if csv_files:
            latest_csv = os.path.join(output_dir, max(csv_files, key=lambda x: os.path.getctime(os.path.join(output_dir, x))))
            return latest_csv
    
    # Search in current folder
    csv_files = [f for f in os.listdir('.') if f.startswith('cloudtrail_log_') and f.endswith('.csv')]
    if csv_files:
        return max(csv_files, key=os.path.getctime)
    
    return None


def run_cloudtrail_parser():
    """Function to run CloudTrail parser"""
    print("🔄 Starting CloudTrail log parsing...")
    
    try:
        # Import and run CloudTrail parser
        from cloudtrail_log_parser import CloudTrailParser
        
        # Set default path (user can modify)
        input_path = r"C:\Users\kelly.jang\Desktop\AWS_LOG\CloudTrail_Log_S3"
        
        # Create directory structure
        result_dir = "result"
        parse_log_dir = os.path.join(result_dir, "parse_log")
        
        # Create output directory
        os.makedirs(parse_log_dir, exist_ok=True)
        
        # Run parser
        parser = CloudTrailParser()
        events = parser.parse_directory(input_path)
        
        if events:
            # Generate output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(parse_log_dir, f"cloudtrail_log_{timestamp}.csv")
            
            # Save CSV
            parser.save_to_csv(events, output_file)
            print(f"[SUCCESS] CloudTrail parsing completed: {output_file}")
            return output_file
        else:
            print("[ERROR] No events parsed.")
            return None
            
    except Exception as e:
        print(f"[ERROR] Error during CloudTrail parsing: {e}")
        return None


def main():
    """Main function"""
    print("[INFO] CloudTrail Log Analyzer")
    print("=" * 50)
    
    # Create directory structure
    result_dir = "result"
    parse_log_dir = os.path.join(result_dir, "parse_log")
    analysis_dir = os.path.join(result_dir, "analysis")
    report_dir = os.path.join(result_dir, "report")
    
    # Create directories
    os.makedirs(parse_log_dir, exist_ok=True)
    os.makedirs(analysis_dir, exist_ok=True)
    os.makedirs(report_dir, exist_ok=True)
    
    # 1. Find CloudTrail CSV file
    latest_csv = find_cloudtrail_csv()
    
    if not latest_csv:
        print("📁 CloudTrail CSV file not found.")
        print("🔄 Starting CloudTrail log parsing...")
        
        # Run CloudTrail parser
        latest_csv = run_cloudtrail_parser()
        
        if not latest_csv:
            print("[ERROR] CloudTrail parsing failed.")
            return
    else:
        print(f"📁 Analysis file: {latest_csv}")
    
    # 2. Create and run analyzer
    print(f"\n[INFO] Starting CloudTrail log analysis...")
    analyzer = CloudTrailAnalyzer(latest_csv)
    
    # 3. Generate output filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(analysis_dir, f"cloudtrail_analysis_{timestamp}.xlsx")
    
    # 4. Create Excel file and collect analysis data
    analysis_data = analyzer.save_analysis_to_excel(output_file)
    
    # 5. Generate HTML report (reuse analysis data)
    html_file = analyzer.generate_html_report(output_file, analysis_data)
    
    print(f"\n[COMPLETE] CloudTrail log analysis completed!")
    print(f"[INFO] Excel file: {output_file}")
    print(f"[INFO] HTML report: {html_file}")
    print(f"[INFO] Check the detailed analysis results with 8 sheets and HTML report!")
    
    return output_file, html_file


def run_analysis(csv_file: str, analysis_dir: str, report_dir: str) -> tuple:
    """Run complete CloudTrail analysis and return file paths"""
    try:
        # Create directories
        os.makedirs(analysis_dir, exist_ok=True)
        os.makedirs(report_dir, exist_ok=True)
        
        # Create analyzer instance with CSV file path
        analyzer = CloudTrailAnalyzer(csv_file)
        
        # Generate output filename for Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"cloudtrail_analysis_{timestamp}.xlsx"
        excel_file = os.path.join(analysis_dir, excel_filename)
        
        # Run analysis and save to Excel
        analysis_data = analyzer.save_analysis_to_excel(excel_file)
        
        # Generate HTML report
        html_file = analyzer.generate_html_report(report_dir, analysis_data)
        
        return excel_file, html_file
        
    except Exception as e:
        print(f"[ERROR] CloudTrail analysis failed: {str(e)}")
        raise


if __name__ == "__main__":
    main()